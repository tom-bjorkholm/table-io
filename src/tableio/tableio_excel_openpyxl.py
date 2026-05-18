#! /usr/bin/env python3
"""TableIO reader/writer class for Excel files using OpenPyXL."""

# Copyright (c) 2026 Tom Björkholm
# MIT License

from pathlib import Path
from typing import Callable, Optional
from xml.etree import ElementTree as ET
from zipfile import ZipFile, ZipInfo
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Border, Font, PatternFill, Side
from openpyxl.utils.cell import get_column_letter, range_boundaries
from openpyxl.worksheet.table import Table
from openpyxl.worksheet.worksheet import Worksheet
from mformat.mformat import PathLike
from tableio._archive_rewrite import rewrite_zip_archive, \
    temporary_output_path
from tableio.border_helper import BorderWeight, CellBorder
from tableio.color import Color
from tableio.tableio import Descriptor, FileAccess
from tableio.tableio_excelbased import TableIOExcelBased
from tableio.value_type import Fmt, Value, get_checked_type
from tableio.capability import Capabilities, CAP_ALL_IMPLEMENTED

_HIGHLIGHT_RGB: dict[Color, str] = {
    Color.RED: 'FFFFC7CE',
    Color.GREEN: 'FFC6EFCE',
    Color.YELLOW: 'FFFFFF00'
}
_SPREADSHEET_NS = 'http://schemas.openxmlformats.org/spreadsheetml/2006/main'
_REL_NS = 'http://schemas.openxmlformats.org/package/2006/relationships'
_CONTENT_TYPES_NS = (
    'http://schemas.openxmlformats.org/package/2006/content-types'
)
_XML_NS = {'main': _SPREADSHEET_NS}
_REL_XML_NS = {'rel': _REL_NS}
_CONTENT_TYPES_XML_NS = {'ct': _CONTENT_TYPES_NS}
_SHARED_STRINGS_PATH = 'xl/sharedStrings.xml'
_WORKBOOK_RELS_PATH = 'xl/_rels/workbook.xml.rels'
_CONTENT_TYPES_PATH = '[Content_Types].xml'
_SHARED_STRINGS_REL_TYPE = (
    'http://schemas.openxmlformats.org/officeDocument/2006/'
    'relationships/sharedStrings'
)
_SHARED_STRINGS_CONTENT_TYPE = (
    'application/vnd.openxmlformats-officedocument.'
    'spreadsheetml.sharedStrings+xml'
)
_FONT_CHILD_ORDER = [
    'b',
    'i',
    'u',
    'strike',
    'outline',
    'shadow',
    'condense',
    'extend',
    'sz',
    'color',
    'name',
    'family',
    'charset',
    'scheme'
]


def _xml_tag(namespace: str, tag_name: str) -> str:
    """Return one fully qualified XML tag name."""
    return f'{{{namespace}}}{tag_name}'


def _font_child_sort_key(element: ET.Element) -> int:
    """Return the schema-order key for one Excel font child element."""
    tag_name = element.tag.split('}', 1)[-1]
    if tag_name in _FONT_CHILD_ORDER:
        return _FONT_CHILD_ORDER.index(tag_name)
    return len(_FONT_CHILD_ORDER)


def _styles_xml_with_sorted_fonts(data: bytes) -> bytes:
    """Return styles XML with font child elements in schema order."""
    root = ET.fromstring(data)
    for font in root.findall('main:fonts/main:font', _XML_NS):
        font[:] = sorted(list(font), key=_font_child_sort_key)
    return bytes(ET.tostring(root, encoding='utf-8', xml_declaration=True))


def _new_shared_strings_root() -> ET.Element:
    """Create an empty shared strings XML root element."""
    return ET.Element(_xml_tag(_SPREADSHEET_NS, 'sst'))


def _read_shared_strings_root(file_name: Path) -> ET.Element:
    """Read the shared strings root, or create an empty one."""
    try:
        with ZipFile(file_name, 'r') as archive:
            data = archive.read(_SHARED_STRINGS_PATH)
    except KeyError:
        return _new_shared_strings_root()
    return ET.fromstring(data)


def _shared_string_count(shared_strings_root: ET.Element) -> int:
    """Return the number of shared string items in the root."""
    return len(shared_strings_root.findall('main:si', _XML_NS))


def _shared_strings_xml(shared_strings_root: ET.Element) -> bytes:
    """Return finalized shared strings XML bytes."""
    count_text = str(_shared_string_count(shared_strings_root))
    shared_strings_root.set('count', count_text)
    shared_strings_root.set('uniqueCount', count_text)
    return bytes(ET.tostring(shared_strings_root, encoding='utf-8',
                             xml_declaration=True))


def _inline_string_to_shared_string(cell: ET.Element,
                                    shared_strings_root: ET.Element) -> None:
    """Move one inline string cell value to the shared string table."""
    inline_string = cell.find('main:is', _XML_NS)
    if inline_string is None:
        return
    shared_index = _shared_string_count(shared_strings_root)
    shared_item = ET.Element(_xml_tag(_SPREADSHEET_NS, 'si'))
    for child in list(inline_string):
        inline_string.remove(child)
        shared_item.append(child)
    shared_strings_root.append(shared_item)
    cell.remove(inline_string)
    cell.set('t', 's')
    value = ET.Element(_xml_tag(_SPREADSHEET_NS, 'v'))
    value.text = str(shared_index)
    cell.append(value)


def _sheet_xml_with_shared_strings(data: bytes,
                                   shared_strings_root: ET.Element) -> bytes:
    """Return sheet XML with inline strings converted to shared strings."""
    root = ET.fromstring(data)
    for cell in root.findall('.//main:c[@t="inlineStr"]', _XML_NS):
        _inline_string_to_shared_string(cell, shared_strings_root)
    return bytes(ET.tostring(root, encoding='utf-8', xml_declaration=True))


def _content_types_with_shared_strings(data: bytes) -> bytes:
    """Return content types XML with a shared strings override."""
    root = ET.fromstring(data)
    for override in root.findall('ct:Override', _CONTENT_TYPES_XML_NS):
        if override.get('PartName') == '/' + _SHARED_STRINGS_PATH:
            return data
    override = ET.Element(_xml_tag(_CONTENT_TYPES_NS, 'Override'))
    override.set('PartName', '/' + _SHARED_STRINGS_PATH)
    override.set('ContentType', _SHARED_STRINGS_CONTENT_TYPE)
    root.append(override)
    return bytes(ET.tostring(root, encoding='utf-8', xml_declaration=True))


def _next_relationship_id(root: ET.Element) -> str:
    """Return the next workbook relationship id."""
    max_id = 0
    for relationship in root.findall('rel:Relationship', _REL_XML_NS):
        rel_id = relationship.get('Id', '')
        if not rel_id.startswith('rId'):
            continue
        try:
            max_id = max(max_id, int(rel_id[3:]))
        except ValueError:
            continue
    return f'rId{max_id + 1}'


def _workbook_rels_with_shared_strings(data: bytes) -> bytes:
    """Return workbook relationships XML with a shared strings relation."""
    root = ET.fromstring(data)
    for relationship in root.findall('rel:Relationship', _REL_XML_NS):
        if relationship.get('Type') == _SHARED_STRINGS_REL_TYPE:
            return data
    relationship = ET.Element(_xml_tag(_REL_NS, 'Relationship'))
    relationship.set('Id', _next_relationship_id(root))
    relationship.set('Type', _SHARED_STRINGS_REL_TYPE)
    relationship.set('Target', 'sharedStrings.xml')
    root.append(relationship)
    return bytes(ET.tostring(root, encoding='utf-8', xml_declaration=True))


def _is_worksheet_xml(filename: str) -> bool:
    """Return True if an archive entry is a worksheet XML file."""
    return filename.startswith('xl/worksheets/') and filename.endswith('.xml')


def _worksheet_rewrites_with_shared_strings(
        file_name: Path, shared_strings_root: ET.Element) -> dict[str, bytes]:
    """Return worksheet XML rewrites and update the shared string table."""
    rewrites: dict[str, bytes] = {}
    with ZipFile(file_name, 'r') as archive:
        for filename in archive.namelist():
            if not _is_worksheet_xml(filename):
                continue
            data = archive.read(filename)
            if b'inlineStr' not in data:
                continue
            rewrites[filename] = _sheet_xml_with_shared_strings(
                data, shared_strings_root)
    return rewrites


def _rewrite_saved_workbook(file_name: Path) -> None:
    """Rewrite the saved workbook so styles XML follows validator order."""
    shared_strings_root = _read_shared_strings_root(file_name)
    worksheet_rewrites = _worksheet_rewrites_with_shared_strings(
        file_name, shared_strings_root)
    use_shared_strings = _shared_string_count(shared_strings_root) > 0

    def rewrite_entry(item: ZipInfo, data: bytes) -> Optional[bytes]:
        """Rewrite one workbook archive entry when needed."""
        filename = item.filename
        if filename == 'xl/styles.xml':
            return _styles_xml_with_sorted_fonts(data)
        if use_shared_strings and filename == _SHARED_STRINGS_PATH:
            return None
        if use_shared_strings and filename == _CONTENT_TYPES_PATH:
            return _content_types_with_shared_strings(data)
        if use_shared_strings and filename == _WORKBOOK_RELS_PATH:
            return _workbook_rels_with_shared_strings(data)
        if filename in worksheet_rewrites:
            return worksheet_rewrites[filename]
        return data
    extra_entries = None
    if use_shared_strings:
        extra_entries = {
            _SHARED_STRINGS_PATH: _shared_strings_xml(shared_strings_root)
        }
    rewrite_zip_archive(file_name, rewrite_entry, extra_entries=extra_entries)


class TableIOExcelOpenPyXL(TableIOExcelBased):
    """TableIO reader/writer class for Excel files using OpenPyXL.

    The implementation operates on one current worksheet at a time. In
    UPDATE mode the default write position is after the last used row in
    the selected worksheet.
    """

    def __init__(self, file_name: PathLike, file_access: FileAccess,
                 file_exists_callback: Optional[Callable[[str], None]]
                 = None):
        """Initialize the TableIOExcelOpenPyXL class."""
        super().__init__(file_name=file_name, file_access=file_access,
                         file_exists_callback=file_exists_callback)
        self.workbook: Optional[Workbook] = None
        self.read_workbook: Optional[Workbook] = None
        self.worksheet: Optional[Worksheet] = None
        self.read_worksheet: Optional[Worksheet] = None
        self._border_cache: dict[CellBorder, Border] = {}

    @classmethod
    def get_capabilities(cls) -> Capabilities:
        """Return the standard spreadsheet backend capabilities."""
        return CAP_ALL_IMPLEMENTED

    @classmethod
    def get_description(cls) -> Descriptor:
        """Get the description of the TableIOExcelOpenPyXL class."""
        return Descriptor(format_name='Excel', implementation='OpenPyXL',
                          capabilities=cls.get_capabilities(),
                          mandatory_args=[], optional_args=[])

    def open(self) -> None:
        """Open the Excel workbook."""
        if self.workbook is not None:
            raise RuntimeError(f'File {self.file_name} already open')
        self._border_cache = {}
        if self.file_access == FileAccess.CREATE:
            workbook = Workbook()
            self.workbook = workbook
            self.read_workbook = workbook
        elif self.file_access == FileAccess.READ:
            workbook = load_workbook(self.file_name, data_only=True)
            self.workbook = workbook
            self.read_workbook = workbook
        else:
            self.workbook = load_workbook(self.file_name)
            self.read_workbook = load_workbook(self.file_name, data_only=True)
        assert self.workbook is not None
        assert self.read_workbook is not None
        worksheet = self.workbook.active
        read_worksheet = self.read_workbook.active
        assert isinstance(worksheet, Worksheet)
        assert isinstance(read_worksheet, Worksheet)
        self.worksheet = worksheet
        self.read_worksheet = read_worksheet
        self._initialize_positions()

    def _end_state(self) -> None:
        """Finalize in-memory state before closing."""

    def _write_file_suffix(self) -> None:
        """Write the workbook to disk when the file is writable."""
        if self.file_access == FileAccess.READ:
            return
        assert self.workbook is not None
        file_path = Path(self.file_name)
        temp_path = temporary_output_path(file_path, '.xlsx')
        try:
            self.workbook.save(temp_path)
            _rewrite_saved_workbook(temp_path)
            temp_path.replace(file_path)
        finally:
            if temp_path.exists():
                temp_path.unlink()

    def _close(self) -> None:
        """Close any open workbook handles."""
        if self.read_workbook is not None and \
                self.read_workbook is not self.workbook:
            self.read_workbook.close()
            self.read_workbook = None
            self.read_worksheet = None
        if self.workbook is not None:
            self.workbook.close()
            self.workbook = None
            self.worksheet = None
        self.read_workbook = None
        self.read_worksheet = None

    @staticmethod
    def _worksheet_name_map(workbook: Workbook) -> dict[str, Worksheet]:
        """Return the workbook worksheets indexed case-insensitively."""
        ret: dict[str, Worksheet] = {}
        for worksheet in workbook.worksheets:
            ret[worksheet.title.casefold()] = worksheet
        return ret

    def _set_active_worksheets(self, worksheet: Worksheet,
                               read_worksheet: Worksheet) -> None:
        """Set the current writable and readable worksheets."""
        assert self.workbook is not None
        assert self.read_workbook is not None
        self.worksheet = worksheet
        self.read_worksheet = read_worksheet
        self.workbook.active = self.workbook.index(worksheet)
        if self.read_workbook is self.workbook:
            return
        self.read_workbook.active = self.read_workbook.index(read_worksheet)

    def _list_sheets(self) -> list[str]:
        """List the sheets in the workbook."""
        assert self.workbook is not None
        return list(self.workbook.sheetnames)

    def _select_sheet(self, sheet_name: str, create: bool = False) -> None:
        """Select one workbook sheet, optionally creating it."""
        assert self.workbook is not None
        assert self.read_workbook is not None
        assert self.worksheet is not None
        assert self.read_worksheet is not None
        sheet_key = self._sheet_key(sheet_name)
        workbook_map = self._worksheet_name_map(self.workbook)
        read_map = self._worksheet_name_map(self.read_workbook)
        worksheet = workbook_map.get(sheet_key)
        read_worksheet = read_map.get(sheet_key)
        if worksheet is None:
            if not create:
                raise KeyError(sheet_name)
            self._check_file_is_writable()
            worksheet = self.workbook.create_sheet(title=sheet_name)
            if self.read_workbook is self.workbook:
                read_worksheet = worksheet
            else:
                read_worksheet = self.read_workbook.create_sheet(
                    title=sheet_name)
        assert read_worksheet is not None
        self._save_current_sheet_state()
        self._set_active_worksheets(worksheet, read_worksheet)
        self._load_current_sheet_state()

    def _current_sheet_name(self) -> str:
        """Return the name of the selected worksheet."""
        assert self.worksheet is not None
        return self.worksheet.title

    def _read_sheet(self) -> object:
        """Return the readable worksheet."""
        assert self.read_worksheet is not None
        return self.read_worksheet

    def _write_sheet(self) -> object:
        """Return the writable worksheet."""
        assert self.worksheet is not None
        return self.worksheet

    @staticmethod
    def _highlight_fill(highlight: Color) -> PatternFill:
        """Return the fill object for the requested highlight color."""
        if highlight == Color.NONE:
            return PatternFill()
        return PatternFill(fill_type='solid',
                           fgColor=_HIGHLIGHT_RGB[highlight])

    def _write_value_to_sheet(self, sheet: object, row: int, column: int,
                              value: object) -> None:
        """Write one value to one worksheet cell."""
        worksheet = get_checked_type(sheet, Worksheet)
        cell = worksheet.cell(row=row + 1, column=column + 1)
        cell.style = 'Normal'
        cell.value = self._spreadsheet_value_from_python(value)

    def _set_cell_format(self, sheet: object, row: int, column: int,
                         fmt: Optional[Fmt]) -> None:
        """Apply cell formatting to one worksheet cell."""
        if fmt is None:
            return
        worksheet = get_checked_type(sheet, Worksheet)
        cell = worksheet.cell(row=row + 1, column=column + 1)
        cell.font = Font(bold=fmt.bold, italic=fmt.italic)
        cell.fill = self._highlight_fill(fmt.highlight)

    @staticmethod
    def _border_side(weight: BorderWeight) -> Side:
        """Return one OpenPyXL side object for the requested weight."""
        if weight == BorderWeight.NONE:
            return Side()
        if weight == BorderWeight.THIN:
            return Side(style='thin')
        return Side(style='medium')

    def _border_value(self, borders: CellBorder) -> Border:
        """Return one cached OpenPyXL border object."""
        cached = self._border_cache.get(borders)
        if cached is not None:
            return cached
        border = Border(left=self._border_side(borders.left),
                        right=self._border_side(borders.right),
                        top=self._border_side(borders.top),
                        bottom=self._border_side(borders.bottom))
        self._border_cache[borders] = border
        return border

    def _set_cell_borders(self, sheet: object, row: int, column: int,
                          borders: CellBorder) -> None:
        """Apply normalized borders to one worksheet cell."""
        worksheet = get_checked_type(sheet, Worksheet)
        cell = worksheet.cell(row=row + 1, column=column + 1)
        cell.border = self._border_value(borders)

    def _apply_heading_style(self, row: int, column: int, level: int) -> None:
        """Apply the heading font to one worksheet cell."""
        assert self.worksheet is not None
        cell = self.worksheet.cell(row=row + 1, column=column + 1)
        cell.font = Font(bold=True, size=self._heading_font_size(level))

    def _last_used_row(self, sheet: object) -> int:
        """Return the last used row index on a worksheet."""
        worksheet = get_checked_type(sheet, Worksheet)
        return self._used_bounds_by_cell_scan(worksheet, worksheet.max_row,
                                              worksheet.max_column)[0]

    def _last_used_column(self, sheet: object) -> int:
        """Return the last used column index on a worksheet."""
        worksheet = get_checked_type(sheet, Worksheet)
        return self._used_bounds_by_cell_scan(worksheet, worksheet.max_row,
                                              worksheet.max_column)[1]

    def _cell_value(self, sheet: object, row: int, column: int) -> Value:
        """Return one worksheet cell as a public Value."""
        worksheet = get_checked_type(sheet, Worksheet)
        raw = worksheet.cell(row=row + 1, column=column + 1).value
        return self._python_value_from_spreadsheet(raw)

    @staticmethod
    def _table_bounds(table_ref: str) -> tuple[int, int, int, int]:
        """Return zero-based exclusive bounds for one worksheet table."""
        min_column, min_row, max_column, max_row = range_boundaries(table_ref)
        assert min_column is not None
        assert min_row is not None
        assert max_column is not None
        assert max_row is not None
        return min_row - 1, min_column - 1, max_row, max_column

    def _filtered_range_infos(self) -> list[tuple[str, tuple[int, int,
                                                             int, int]]]:
        """Return the worksheet tables and their bounds."""
        assert self.worksheet is not None
        ret: list[tuple[str, tuple[int, int, int, int]]] = []
        for table_name in list(self.worksheet.tables):
            table = self.worksheet.tables[table_name]
            ret.append((table_name, self._table_bounds(table.ref)))
        return ret

    def _delete_filtered_range(self, name: str) -> None:
        """Delete one worksheet table by name."""
        assert self.worksheet is not None
        del self.worksheet.tables[name]

    def _normalize_filtered_table_header(self, top: int, left: int,
                                         right: int) -> None:
        """Convert the filtered table header row to strings when needed."""
        assert self.worksheet is not None
        assert self.read_worksheet is not None
        headers = self._filtered_table_headers([
            self.worksheet.cell(row=top + 1, column=column + 1).value
            for column in range(left, right)
        ])
        for column_offset, header_value in enumerate(headers):
            column = left + column_offset
            cell = self.worksheet.cell(row=top + 1, column=column + 1)
            cell.value = header_value
            if self.read_worksheet is not self.worksheet:
                read_cell = self.read_worksheet.cell(row=top + 1,
                                                     column=column + 1)
                read_cell.value = header_value

    def _add_filtered_range(self, bounds: tuple[int, int, int, int],
                            name: str) -> None:
        """Add one lightweight Excel table for a filtered data range."""
        assert self.worksheet is not None
        self._normalize_filtered_table_header(bounds[0], bounds[1], bounds[3])
        self.worksheet.add_table(
            Table(displayName=name,
                  ref=self._excel_range_ref(bounds[0], bounds[1], bounds[2],
                                            bounds[3])))

    def _set_column_width_if_wider(self, column: int, width: float) -> None:
        """Widen one worksheet column if the target width is larger."""
        assert self.worksheet is not None
        column_letter = get_column_letter(column + 1)
        column_dimension = self.worksheet.column_dimensions[column_letter]
        if column_dimension.width is None or width > column_dimension.width:
            column_dimension.width = width
