#! /usr/bin/env python3
"""Spreadsheet validation helpers for example tests."""

# Copyright (c) 2026 Tom Björkholm
# MIT License

from datetime import date, datetime, time
from functools import partial
from pathlib import Path
from typing import Callable, NamedTuple, Optional, cast
from odfdo import Document, Table
from odfdo.style import Style
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from openxml_audit import (  # type: ignore[import-untyped]
    OdfValidator,
    OpenXmlValidator,
    ValidationResult
)
from python_calamine import (
    CalamineSheet,
    CalamineWorkbook,
    load_workbook as load_calamine_workbook
)
from tableio.border_helper import BorderHelper, BorderWeight, CellBorder, \
    NO_BORDERS
from tableio.capability import CAP_NEEDED, Capabilities
from tableio.color import Color
from tableio.tableio_types import TableBorderStyle
from tableio.value_type import Value, get_checked_type


class ExpectedCellStyle(NamedTuple):
    """Expected style values for one or more spreadsheet cells.

    Each non-``None`` field must match in every checked cell. A value of
    ``None`` means that the corresponding style aspect is not checked.
    """

    bold: Optional[bool] = None
    italic: Optional[bool] = None
    background_color: Optional[Color] = None


class RelativeStyleExpectation(NamedTuple):
    """Expected style for a rectangular area relative to an anchor cell.

    The anchor cell is at offset ``(0, 0)``. ``row_offset`` and
    ``col_offset`` point to the top-left cell of the checked area relative
    to the anchor cell. ``number_of_rows`` and ``number_of_columns`` define
    the size of that area and must both be positive. Omitting the offsets
    and size checks the anchor cell itself.
    """

    expected_style: ExpectedCellStyle
    row_offset: int = 0
    col_offset: int = 0
    number_of_rows: int = 1
    number_of_columns: int = 1


class AnchoredStyleExpectation(NamedTuple):
    """Style checks anchored on the first matching row fragment.

    The row fragment is matched on the specified sheet as a left-to-right
    ordered subsequence within one row. For example ``['hello', 'world']``
    matches the row ``['hello', 'wonderful', 'world']``. The anchor cell is
    the first cell of the first match, selected by the lowest row number and
    then the lowest column number. Each relative expectation checks one cell
    or one rectangular area starting from that anchor.
    """

    sheet_name: str
    anchor_row_fragment: list[Value]
    relative_expectations: list[RelativeStyleExpectation]


class RelativeBorderExpectation(NamedTuple):
    """Expected border style for one rectangular area relative to an anchor.

    The checked rectangular area is interpreted as one written table with the
    given size. Each cell border in that area must match the normalized cell
    borders implied by ``border_style``.
    """

    border_style: TableBorderStyle
    row_offset: int = 0
    col_offset: int = 0
    number_of_rows: int = 1
    number_of_columns: int = 1


class AnchoredBorderExpectation(NamedTuple):
    """Border checks anchored on the first matching row fragment.

    The row fragment is matched like in ``AnchoredStyleExpectation``. The
    anchor cell is the first cell of the first match, and each relative
    border expectation checks one rectangular area starting from that anchor.
    """

    sheet_name: str
    anchor_row_fragment: list[Value]
    relative_expectations: list[RelativeBorderExpectation]


class SheetContentExpectation(NamedTuple):
    """Ordered row-fragment expectations for one sheet.

    Each row fragment is matched as a left-to-right ordered subsequence
    within one row. For example ``['hello', 'world']`` matches the row
    ``['hello', 'wonderful', 'world']``.
    """

    sheet_name: str
    row_fragments: list[list[Value]]


_BORDER_CHECK_CAPABILITIES = Capabilities(can_write_borders=CAP_NEEDED)


def _spreadsheet_output_path(output_base: Path | str, suffix: str) -> Path:
    """Return the actual file path after tableio appends its suffix."""
    normalized_suffix = suffix if suffix.startswith('.') else f'.{suffix}'
    return Path(f'{output_base}{normalized_suffix}')


def _normalize_file_path(file_name: Path | str) -> Path:
    """Return one normalized file path."""
    return Path(file_name)


def _spreadsheet_suffix(file_path: Path) -> str:
    """Return the supported spreadsheet suffix in lowercase."""
    suffix = file_path.suffix.lower()
    if suffix not in ['.xlsx', '.ods']:
        raise ValueError(f'Unsupported spreadsheet suffix: {file_path}')
    return suffix


def _check_file_exists(file_path: Path) -> None:
    """Check that one target file exists."""
    if not file_path.is_file():
        raise AssertionError(f'Spreadsheet file does not exist: {file_path}')


def _syntax_validation_result(file_path: Path) -> ValidationResult:
    """Return the syntax-validation result for one spreadsheet file."""
    suffix = _spreadsheet_suffix(file_path)
    if suffix == '.xlsx':
        return OpenXmlValidator().validate(file_path)
    return OdfValidator().validate(file_path)


def _raise_syntax_error(file_path: Path,
                        validation_result: ValidationResult) -> None:
    """Raise one assertion with collected syntax-validation errors."""
    formatted_errors = '\n'.join(
        str(error) for error in validation_result.errors[:20]
    )
    raise AssertionError(
        f'Syntax validation failed for {file_path}:\n{formatted_errors}'
    )


def _match_expected_error(actual_errors: list[str],
                          used_indexes: set[int],
                          expected_error: str) -> Optional[int]:
    """Return the index of the first unmatched actual error containing text."""
    for index, actual_error in enumerate(actual_errors):
        if index in used_indexes:
            continue
        if expected_error in actual_error:
            return index
    return None


def _check_expected_syntax_errors(
        file_path: Path,
        validation_result: ValidationResult,
        expected_errors: Optional[list[str]]) -> None:
    """Check actual syntax errors against optional expected substrings."""
    actual_errors = [
        str(error) for error in validation_result.errors
    ]
    if expected_errors is None:
        if actual_errors:
            _raise_syntax_error(file_path, validation_result)
        return
    used_indexes: set[int] = set()
    missing_expected: list[str] = []
    for expected_error in expected_errors:
        index = _match_expected_error(actual_errors, used_indexes,
                                      expected_error)
        if index is None:
            missing_expected.append(expected_error)
            continue
        used_indexes.add(index)
    unexpected_errors = [
        error for index, error in enumerate(actual_errors)
        if index not in used_indexes
    ]
    if not missing_expected and not unexpected_errors:
        return
    formatted_expected = '\n'.join(missing_expected[:20]) or '(none)'
    formatted_unexpected = '\n'.join(unexpected_errors[:20]) or '(none)'
    raise AssertionError(
        f'Syntax validation mismatch for {file_path}:\n'
        f'Missing expected errors:\n{formatted_expected}\n'
        f'Unexpected errors:\n{formatted_unexpected}'
    )


def _normalize_cell_value(value: object) -> Value:
    """Return one spreadsheet cell value normalized to one public Value."""
    if value is None:
        return ''
    if isinstance(value, datetime):
        return value
    if isinstance(value, date):
        return datetime.combine(value, time())
    if isinstance(value, (str, bool, int, float)):
        return cast(Value, value)
    return str(value)


def _sheet_rows(sheet: CalamineSheet) -> list[list[Value]]:
    """Return one sheet as normalized rows of Value cells."""
    rows = cast(list[list[object]],
                sheet.to_python(skip_empty_area=False))
    return [
        [_normalize_cell_value(value) for value in row]
        for row in rows
    ]


def _read_spreadsheet_content(file_path: Path) -> dict[str, list[list[Value]]]:
    """Read one spreadsheet into typed rows keyed by sheet name."""
    workbook: CalamineWorkbook = load_calamine_workbook(file_path)
    try:
        return {
            sheet_name: _sheet_rows(workbook.get_sheet_by_name(sheet_name))
            for sheet_name in workbook.sheet_names
        }
    finally:
        workbook.close()


def _match_row_fragment(row_values: list[Value],
                        row_fragment: list[Value]) -> Optional[int]:
    """Return the first matching column for one row fragment."""
    if not row_fragment:
        raise AssertionError('Row fragments must not be empty.')
    for start_column, cell_value in enumerate(row_values):
        if cell_value != row_fragment[0]:
            continue
        next_fragment_index = 1
        for row_value in row_values[start_column + 1:]:
            if next_fragment_index == len(row_fragment):
                break
            if row_value == row_fragment[next_fragment_index]:
                next_fragment_index += 1
        if next_fragment_index == len(row_fragment):
            return start_column
    return None


def _find_matching_row(rows: list[list[Value]],
                       row_fragment: list[Value],
                       start_row: int = 0) -> Optional[tuple[int, int]]:
    """Return the first row and column matching one row fragment."""
    for row_index in range(start_row, len(rows)):
        start_column = _match_row_fragment(rows[row_index], row_fragment)
        if start_column is not None:
            return row_index, start_column
    return None


def _check_sheet_names(actual_sheet_names: list[str],
                       expected_fragments:
                       list[SheetContentExpectation]) -> None:
    """Check that actual sheet names match the expected sheet order."""
    expected_sheet_names = [
        fragment.sheet_name for fragment in expected_fragments
    ]
    if actual_sheet_names != expected_sheet_names:
        raise AssertionError(
            'Unexpected sheet names. '
            f'Expected {expected_sheet_names}, got {actual_sheet_names}.'
        )


def _ods_table(document: Document, sheet_name: str) -> Table:
    """Return one ODS table by sheet name."""
    table = document.body.get_table(name=sheet_name)
    if table is None:
        raise AssertionError(f'Missing sheet {sheet_name!r} in ODS document.')
    return get_checked_type(table, Table)


def _border_weight_text(weight: BorderWeight) -> str:
    """Return one readable border-weight name."""
    return weight.name.lower()


def _color_from_rgb_text(rgb_text: Optional[str]) -> Color:
    """Return one tableio color from a workbook RGB string."""
    if rgb_text is None:
        return Color.NONE
    normalized = rgb_text.upper()
    if normalized.startswith('#'):
        normalized = normalized[1:]
    normalized = normalized[-6:]
    color_map = {
        'FFC7CE': Color.RED,
        'C6EFCE': Color.GREEN,
        'FFFF00': Color.YELLOW
    }
    return color_map.get(normalized, Color.NONE)


def _excel_border_weight(style_name: Optional[str]) -> BorderWeight:
    """Return one normalized border weight from one Excel side style."""
    if style_name in [None, 'none']:
        return BorderWeight.NONE
    if style_name in ['hair', 'dashDot', 'dashDotDot', 'dashed',
                      'dotted', 'slantDashDot', 'thin']:
        return BorderWeight.THIN
    if style_name in ['double', 'medium', 'mediumDashDot',
                      'mediumDashDotDot', 'mediumDashed', 'thick']:
        return BorderWeight.THICK
    raise ValueError(f'Unsupported Excel border style: {style_name}')


def _excel_fill_color(worksheet: Worksheet,
                      row_index: int,
                      col_index: int) -> Color:
    """Return one Excel cell background color."""
    cell = worksheet.cell(row=row_index + 1, column=col_index + 1)
    if cell.fill.fill_type != 'solid':
        return Color.NONE
    rgb_value = cell.fill.fgColor.rgb
    if rgb_value is None:
        rgb_value = cell.fill.start_color.rgb
    return _color_from_rgb_text(rgb_value)


def _excel_actual_borders(worksheet: Worksheet,
                          row_index: int,
                          col_index: int) -> CellBorder:
    """Return actual border weights for one Excel cell."""
    cell = worksheet.cell(row=row_index + 1, column=col_index + 1)
    return CellBorder(
        top=_excel_border_weight(cell.border.top.style),
        right=_excel_border_weight(cell.border.right.style),
        bottom=_excel_border_weight(cell.border.bottom.style),
        left=_excel_border_weight(cell.border.left.style)
    )


def _excel_actual_style(worksheet: Worksheet,
                        row_index: int,
                        col_index: int) -> ExpectedCellStyle:
    """Return actual style values for one Excel cell."""
    cell = worksheet.cell(row=row_index + 1, column=col_index + 1)
    return ExpectedCellStyle(
        bold=bool(cell.font.bold),
        italic=bool(cell.font.italic),
        background_color=_excel_fill_color(worksheet, row_index, col_index)
    )


def _ods_border_weight(border_text: Optional[str]) -> BorderWeight:
    """Return one normalized border weight from one ODS border string."""
    if border_text is None or border_text == '':
        return BorderWeight.NONE
    size_text = border_text.split()[0]
    if not size_text.endswith('pt'):
        raise ValueError(f'Unsupported ODS border width: {border_text}')
    try:
        width = float(size_text[:-2])
    except ValueError as exc:
        raise ValueError(
            f'Unsupported ODS border width: {border_text}'
        ) from exc
    if width < 1.0:
        return BorderWeight.THIN
    return BorderWeight.THICK


def _ods_actual_borders(document: Document,
                        table: Table,
                        row_index: int,
                        col_index: int) -> CellBorder:
    """Return actual border weights for one ODS cell."""
    cell = table.get_cell((col_index, row_index), clone=False)
    if cell.style is None:
        return NO_BORDERS
    style = document.get_style('table-cell', cell.style)
    if style is None:
        return NO_BORDERS
    checked_style = get_checked_type(style, Style)
    table_props = cast(
        dict[str, str], checked_style.get_properties('table-cell') or {}
    )
    return CellBorder(
        top=_ods_border_weight(table_props.get('fo:border-top')),
        right=_ods_border_weight(table_props.get('fo:border-right')),
        bottom=_ods_border_weight(table_props.get('fo:border-bottom')),
        left=_ods_border_weight(table_props.get('fo:border-left'))
    )


def _ods_actual_style(document: Document,
                      table: Table,
                      row_index: int,
                      col_index: int) -> ExpectedCellStyle:
    """Return actual style values for one ODS cell."""
    cell = table.get_cell((col_index, row_index), clone=False)
    if cell.style is None:
        return ExpectedCellStyle(
            bold=False,
            italic=False,
            background_color=Color.NONE
        )
    style = document.get_style('table-cell', cell.style)
    if style is None:
        return ExpectedCellStyle(
            bold=False,
            italic=False,
            background_color=Color.NONE
        )
    checked_style = get_checked_type(style, Style)
    table_props = cast(
        dict[str, str], checked_style.get_properties('table-cell') or {}
    )
    text_props = cast(
        dict[str, str], checked_style.get_properties('text') or {}
    )
    return ExpectedCellStyle(
        bold=text_props.get('fo:font-weight') == 'bold',
        italic=text_props.get('fo:font-style') == 'italic',
        background_color=_color_from_rgb_text(
            table_props.get('fo:background-color')
        )
    )


def _cell_position_text(row_index: int, col_index: int) -> str:
    """Return a readable cell position string."""
    return f'row {row_index + 1}, column {col_index + 1}'


def _cell_location_text(file_path: Path,
                        sheet_name: str,
                        row_index: int,
                        col_index: int) -> str:
    """Return a readable file, sheet and cell location string."""
    return (
        f'{_cell_position_text(row_index, col_index)} '
        f'in sheet {sheet_name!r} of {file_path}'
    )


def _area_location_text(file_path: Path,
                        sheet_name: str,
                        top_left: tuple[int, int],
                        number_of_rows: int,
                        number_of_columns: int) -> str:
    """Return a readable checked-area location string."""
    row_index, col_index = top_left
    return (
        f'area starting at {_cell_position_text(row_index, col_index)} '
        f'with size {number_of_rows} x {number_of_columns} '
        f'in sheet {sheet_name!r} of {file_path}'
    )


def _style_mismatch_messages(
        location_text: str,
        expected_style: ExpectedCellStyle,
        actual_style: ExpectedCellStyle) -> list[str]:
    """Return mismatch messages for one actual style."""
    mismatches: list[str] = []
    if (expected_style.bold is not None and
            actual_style.bold != expected_style.bold):
        mismatches.append(
            f'Unexpected bold value at {location_text}. '
            f'Expected {expected_style.bold}, got {actual_style.bold}.'
        )
    if (expected_style.italic is not None and
            actual_style.italic != expected_style.italic):
        mismatches.append(
            f'Unexpected italic value at {location_text}. '
            f'Expected {expected_style.italic}, got {actual_style.italic}.'
        )
    if (expected_style.background_color is not None and
            actual_style.background_color != expected_style.background_color):
        mismatches.append(
            f'Unexpected background color at {location_text}. '
            f'Expected {expected_style.background_color}, '
            f'got {actual_style.background_color}.'
        )
    return mismatches


def _border_mismatch_messages(
        location_text: str,
        expected_borders: CellBorder,
        actual_borders: CellBorder) -> list[str]:
    """Return mismatch messages for one actual cell border."""
    mismatches: list[str] = []
    for side_name, expected_weight, actual_weight in [
            ('top', expected_borders.top, actual_borders.top),
            ('right', expected_borders.right, actual_borders.right),
            ('bottom', expected_borders.bottom, actual_borders.bottom),
            ('left', expected_borders.left, actual_borders.left)]:
        if actual_weight == expected_weight:
            continue
        mismatches.append(
            f'Unexpected {side_name} border at {location_text}. '
            f'Expected {_border_weight_text(expected_weight)}, '
            f'got {_border_weight_text(actual_weight)}.'
        )
    return mismatches


def _checked_area_top_left(anchor_cell: tuple[int, int],
                           row_offset: int,
                           col_offset: int,
                           number_of_rows: int,
                           number_of_columns: int) -> tuple[int, int]:
    """Return the checked area's top-left cell after validation."""
    anchor_row_index, anchor_col_index = anchor_cell
    if number_of_rows <= 0:
        raise ValueError('number_of_rows must be positive.')
    if number_of_columns <= 0:
        raise ValueError('number_of_columns must be positive.')
    target_row_index = anchor_row_index + row_offset
    target_col_index = anchor_col_index + col_offset
    if target_row_index < 0:
        raise ValueError(
            'row_offset points before the first spreadsheet row.'
        )
    if target_col_index < 0:
        raise ValueError(
            'col_offset points before the first spreadsheet column.'
        )
    return target_row_index, target_col_index


def _check_relative_style_expectation(
        file_path: Path,
        sheet_name: str,
        anchor_cell: tuple[int, int],
        relative_expectation: RelativeStyleExpectation,
        actual_style_at:
        Callable[[int, int], ExpectedCellStyle]) -> None:
    """Check one relative style expectation against actual cell styles."""
    target_row_index, target_col_index = _checked_area_top_left(
        anchor_cell,
        relative_expectation.row_offset,
        relative_expectation.col_offset,
        relative_expectation.number_of_rows,
        relative_expectation.number_of_columns
    )
    mismatch_messages: list[str] = []
    for row_index in range(
            target_row_index,
            target_row_index + relative_expectation.number_of_rows):
        for col_index in range(
                target_col_index,
                target_col_index + relative_expectation.number_of_columns):
            actual_style = actual_style_at(row_index, col_index)
            mismatch_messages.extend(
                _style_mismatch_messages(
                    _cell_location_text(
                        file_path, sheet_name, row_index, col_index
                    ),
                    relative_expectation.expected_style,
                    actual_style
                )
            )
    if mismatch_messages:
        area_text = _area_location_text(
            file_path,
            sheet_name,
            (target_row_index, target_col_index),
            relative_expectation.number_of_rows,
            relative_expectation.number_of_columns
        )
        formatted_mismatches = '\n'.join(mismatch_messages)
        raise AssertionError(
            f'Style mismatches in {area_text}:\n'
            f'{formatted_mismatches}'
        )


def _check_relative_border_expectation(  # pylint: disable=too-many-locals
        file_path: Path,
        sheet_name: str,
        anchor_cell: tuple[int, int],
        relative_expectation: RelativeBorderExpectation,
        actual_borders_at:
        Callable[[int, int], CellBorder]) -> None:
    """Check one relative border expectation against actual cell borders."""
    target_row_index, target_col_index = _checked_area_top_left(
        anchor_cell,
        relative_expectation.row_offset,
        relative_expectation.col_offset,
        relative_expectation.number_of_rows,
        relative_expectation.number_of_columns
    )
    border_helper = BorderHelper(relative_expectation.border_style,
                                 _BORDER_CHECK_CAPABILITIES)
    mismatch_messages: list[str] = []
    for row_offset in range(relative_expectation.number_of_rows):
        for col_offset in range(relative_expectation.number_of_columns):
            row_index = target_row_index + row_offset
            col_index = target_col_index + col_offset
            expected_borders = border_helper.cell_border(
                row_offset,
                col_offset,
                relative_expectation.number_of_rows,
                relative_expectation.number_of_columns
            )
            actual_borders = actual_borders_at(row_index, col_index)
            mismatch_messages.extend(
                _border_mismatch_messages(
                    _cell_location_text(
                        file_path, sheet_name, row_index, col_index
                    ),
                    expected_borders,
                    actual_borders
                )
            )
    if mismatch_messages:
        area_text = _area_location_text(
            file_path,
            sheet_name,
            (target_row_index, target_col_index),
            relative_expectation.number_of_rows,
            relative_expectation.number_of_columns
        )
        formatted_mismatches = '\n'.join(mismatch_messages)
        raise AssertionError(
            f'Border mismatches in {area_text}:\n'
            f'{formatted_mismatches}'
        )


def _anchor_match(rows: list[list[Value]],
                  row_fragment: list[Value],
                  sheet_name: str,
                  file_path: Path) -> tuple[int, int]:
    """Return one anchor match or raise a clear assertion."""
    anchor_match = _find_matching_row(rows, row_fragment)
    if anchor_match is None:
        raise AssertionError(
            f'Could not find anchor {row_fragment!r} '
            f'in sheet {sheet_name!r} of {file_path}.'
        )
    return anchor_match


def _check_excel_styles(file_path: Path,
                        workbook_data: dict[str, list[list[Value]]],
                        style_expectations:
                        list[AnchoredStyleExpectation]) -> None:
    """Check style expectations in one Excel workbook."""
    workbook = load_workbook(file_path)
    try:
        for style_expectation in style_expectations:
            rows = workbook_data[style_expectation.sheet_name]
            anchor_row_index, anchor_col_index = _anchor_match(
                rows,
                style_expectation.anchor_row_fragment,
                style_expectation.sheet_name,
                file_path
            )
            worksheet = get_checked_type(
                workbook[style_expectation.sheet_name], Worksheet
            )
            for relative_expectation in (
                    style_expectation.relative_expectations):
                _check_relative_style_expectation(
                    file_path,
                    style_expectation.sheet_name,
                    (anchor_row_index, anchor_col_index),
                    relative_expectation,
                    partial(_excel_actual_style, worksheet)
                )
    finally:
        workbook.close()


def _check_excel_borders(file_path: Path,
                         workbook_data: dict[str, list[list[Value]]],
                         border_expectations:
                         list[AnchoredBorderExpectation]) -> None:
    """Check border expectations in one Excel workbook."""
    workbook = load_workbook(file_path)
    try:
        for border_expectation in border_expectations:
            rows = workbook_data[border_expectation.sheet_name]
            anchor_row_index, anchor_col_index = _anchor_match(
                rows,
                border_expectation.anchor_row_fragment,
                border_expectation.sheet_name,
                file_path
            )
            worksheet = get_checked_type(
                workbook[border_expectation.sheet_name], Worksheet
            )
            for relative_expectation in (
                    border_expectation.relative_expectations):
                _check_relative_border_expectation(
                    file_path,
                    border_expectation.sheet_name,
                    (anchor_row_index, anchor_col_index),
                    relative_expectation,
                    partial(_excel_actual_borders, worksheet)
                )
    finally:
        workbook.close()


def _check_ods_styles(file_path: Path,
                      workbook_data: dict[str, list[list[Value]]],
                      style_expectations:
                      list[AnchoredStyleExpectation]) -> None:
    """Check style expectations in one ODS document."""
    document = Document(file_path)
    for style_expectation in style_expectations:
        rows = workbook_data[style_expectation.sheet_name]
        anchor_row_index, anchor_col_index = _anchor_match(
            rows,
            style_expectation.anchor_row_fragment,
            style_expectation.sheet_name,
            file_path
        )
        table = _ods_table(document, style_expectation.sheet_name)
        for relative_expectation in style_expectation.relative_expectations:
            _check_relative_style_expectation(
                file_path,
                style_expectation.sheet_name,
                (anchor_row_index, anchor_col_index),
                relative_expectation,
                partial(_ods_actual_style, document, table)
            )


def _check_ods_borders(file_path: Path,
                       workbook_data: dict[str, list[list[Value]]],
                       border_expectations:
                       list[AnchoredBorderExpectation]) -> None:
    """Check border expectations in one ODS document."""
    document = Document(file_path)
    for border_expectation in border_expectations:
        rows = workbook_data[border_expectation.sheet_name]
        anchor_row_index, anchor_col_index = _anchor_match(
            rows,
            border_expectation.anchor_row_fragment,
            border_expectation.sheet_name,
            file_path
        )
        table = _ods_table(document, border_expectation.sheet_name)
        for relative_expectation in border_expectation.relative_expectations:
            _check_relative_border_expectation(
                file_path,
                border_expectation.sheet_name,
                (anchor_row_index, anchor_col_index),
                relative_expectation,
                partial(_ods_actual_borders, document, table)
            )


def check_spreadsheet_syntax(
        file_name: Path | str,
        expected_errors: Optional[list[str]] = None) -> None:
    """Check that one spreadsheet file passes syntax validation.

    When ``expected_errors`` is provided, each item is matched as a
    substring against one reported validator error. Every actual error must
    match one expected item, and every expected item must match one actual
    error.
    """
    file_path = _normalize_file_path(file_name)
    _check_file_exists(file_path)
    validation_result = _syntax_validation_result(file_path)
    _check_expected_syntax_errors(file_path, validation_result,
                                  expected_errors)


def check_spreadsheet_content(
        file_name: Path | str,
        expected_fragments: list[SheetContentExpectation]) -> None:
    """Check spreadsheet content against sheet-aware row fragments.

    The actual workbook sheet order must match the order of the
    ``SheetContentExpectation`` items.
    """
    file_path = _normalize_file_path(file_name)
    _check_file_exists(file_path)
    _spreadsheet_suffix(file_path)
    workbook_data = _read_spreadsheet_content(file_path)
    _check_sheet_names(list(workbook_data.keys()), expected_fragments)
    for sheet_expectation in expected_fragments:
        rows = workbook_data[sheet_expectation.sheet_name]
        next_row_index = 0
        for row_fragment in sheet_expectation.row_fragments:
            match = _find_matching_row(rows, row_fragment, next_row_index)
            if match is None:
                raise AssertionError(
                    f'Could not find row fragment {row_fragment!r} '
                    f'in sheet {sheet_expectation.sheet_name!r} '
                    f'of {file_path}.'
                )
            next_row_index = match[0] + 1


def check_spreadsheet_borders(
        file_name: Path | str,
        border_expectations: list[AnchoredBorderExpectation]) -> None:
    """Check spreadsheet borders against anchored border expectations."""
    file_path = _normalize_file_path(file_name)
    _check_file_exists(file_path)
    suffix = _spreadsheet_suffix(file_path)
    if not border_expectations:
        return
    workbook_data = _read_spreadsheet_content(file_path)
    if suffix == '.xlsx':
        _check_excel_borders(file_path, workbook_data, border_expectations)
        return
    _check_ods_borders(file_path, workbook_data, border_expectations)


def check_spreadsheet_styles(
        file_name: Path | str,
        style_expectations: list[AnchoredStyleExpectation]) -> None:
    """Check spreadsheet styles against anchored style expectations.

    Every anchored expectation finds the first matching anchor cell on its
    sheet and then checks each relative expectation from that anchor. A
    relative expectation can target one cell or one rectangular area.
    Cells outside the populated content are treated as unformatted cells if
    the underlying spreadsheet reader reports them that way.
    """
    file_path = _normalize_file_path(file_name)
    _check_file_exists(file_path)
    suffix = _spreadsheet_suffix(file_path)
    if not style_expectations:
        return
    workbook_data = _read_spreadsheet_content(file_path)
    if suffix == '.xlsx':
        _check_excel_styles(file_path, workbook_data, style_expectations)
        return
    _check_ods_styles(file_path, workbook_data, style_expectations)


def check_spreadsheet_file(
        file_name: Path | str,
        expected_fragments: list[SheetContentExpectation],
        style_expectations:
        Optional[list[AnchoredStyleExpectation]] = None,
        border_expectations:
        Optional[list[AnchoredBorderExpectation]] = None,
        expected_errors: Optional[list[str]] = None) -> None:
    """Check spreadsheet syntax, content and optional extra expectations."""
    check_spreadsheet_syntax(file_name, expected_errors)
    check_spreadsheet_content(file_name, expected_fragments)
    if style_expectations is not None:
        check_spreadsheet_styles(file_name, style_expectations)
    if border_expectations is not None:
        check_spreadsheet_borders(file_name, border_expectations)


PLAIN_STYLE: ExpectedCellStyle = \
    ExpectedCellStyle(bold=False, italic=False, background_color=Color.NONE)
BOLD_STYLE: ExpectedCellStyle = \
    ExpectedCellStyle(bold=True, italic=False, background_color=Color.NONE)
ITALIC_STYLE: ExpectedCellStyle = \
    ExpectedCellStyle(bold=False, italic=True, background_color=Color.NONE)
BOLD_ITALIC_STYLE: ExpectedCellStyle = ExpectedCellStyle(bold=True,
                                                         italic=True)
YELLOW_PLAIN: ExpectedCellStyle = \
    ExpectedCellStyle(background_color=Color.YELLOW, italic=False,
                      bold=False)
YELLOW_ITALIC: ExpectedCellStyle = \
    ExpectedCellStyle(background_color=Color.YELLOW, italic=True,
                      bold=False)
YELLOW_BOLD: ExpectedCellStyle = \
    ExpectedCellStyle(background_color=Color.YELLOW, italic=False,
                      bold=True)
YELLOW_BOLD_ITALIC: ExpectedCellStyle = \
    ExpectedCellStyle(background_color=Color.YELLOW, bold=True, italic=True)
RED_PLAIN: ExpectedCellStyle = \
    ExpectedCellStyle(background_color=Color.RED, italic=False,
                      bold=False)
RED_BOLD: ExpectedCellStyle = ExpectedCellStyle(background_color=Color.RED,
                                                bold=True, italic=False)
RED_BOLD_ITALIC: ExpectedCellStyle = \
    ExpectedCellStyle(background_color=Color.RED, bold=True, italic=True)
RED_ITALIC: ExpectedCellStyle = \
    ExpectedCellStyle(background_color=Color.RED, italic=True, bold=False)
GREEN_PLAIN: ExpectedCellStyle = \
    ExpectedCellStyle(background_color=Color.GREEN, italic=False,
                      bold=False)
GREEN_BOLD_ITALIC: ExpectedCellStyle = \
    ExpectedCellStyle(background_color=Color.GREEN, bold=True, italic=True)
GREEN_BOLD: ExpectedCellStyle = ExpectedCellStyle(background_color=Color.GREEN,
                                                  bold=True, italic=False)
GREEN_ITALIC: ExpectedCellStyle = \
    ExpectedCellStyle(background_color=Color.GREEN, italic=True, bold=False)
