#! /usr/bin/env python3
"""Tests for spreadsheet_checkers."""

# Copyright (c) 2026 Tom Björkholm
# MIT License

from pathlib import Path
from typing import Any, Optional
import pytest
from odfdo import Cell, Document, Style, Table
from openpyxl import Workbook as OpenPyXLWorkbook
from openpyxl.styles import Font
from openpyxl.worksheet.worksheet import Worksheet
from xlsxwriter import Workbook  # type: ignore[import-untyped]
from tableio.border_helper import BorderHelper, BorderWeight, CellBorder
from tableio.capability import CAP_NEEDED, Capabilities
from tableio.color import Color
from tableio.tableio_types import TableBorderStyle
from .spreadsheet_checkers import (
    AnchoredBorderExpectation,
    AnchoredStyleExpectation,
    ExpectedCellStyle,
    RelativeBorderExpectation,
    RelativeStyleExpectation,
    SheetContentExpectation,
    check_spreadsheet_borders,
    check_spreadsheet_content,
    check_spreadsheet_file,
    check_spreadsheet_styles,
    check_spreadsheet_syntax,
    _spreadsheet_output_path
)


def _create_basic_xlsx(file_path: Path) -> None:
    """Create one simple `.xlsx` file for checker tests."""
    workbook = Workbook(file_path)
    worksheet = workbook.add_worksheet('Sheet1')
    worksheet.write_row(0, 0, ['hello', 'wonderful', 'world'])
    worksheet.write_string(1, 0, 'later')
    workbook.add_worksheet('Second')
    workbook.close()


def _create_styled_xlsx(file_path: Path) -> None:
    """Create one styled `.xlsx` file for checker tests."""
    workbook = Workbook(file_path)
    worksheet = workbook.add_worksheet('Sheet1')
    yellow_bold_italic = workbook.add_format(
        {'bold': True, 'italic': True, 'bg_color': '#FFFF00'}
    )
    worksheet.write_string(0, 0, 'hello', yellow_bold_italic)
    worksheet.write_string(0, 1, 'wonderful', yellow_bold_italic)
    worksheet.write_string(0, 2, 'world', yellow_bold_italic)
    worksheet.write_string(1, 0, 'hello')
    worksheet.write_string(1, 1, 'world')
    workbook.add_worksheet('Second')
    workbook.close()


def _yellow_ods_style(document: Document) -> str:
    """Insert and return one ODS style with bold italic yellow fill."""
    style = Style('table-cell', name='ce_test_style', area='text',
                  bold=True, italic=True)
    style.set_properties({'fo:background-color': '#ffff00'},
                         area='table-cell')
    document.insert_style(style, automatic=True)
    return 'ce_test_style'


def _create_basic_ods(file_path: Path) -> None:
    """Create one simple `.ods` file for checker tests."""
    document = Document('spreadsheet')
    document.body.clear()
    table = Table('Sheet1')
    document.body.append(table)
    table.set_value((0, 0), 'hello')
    table.set_value((1, 0), 'wonderful')
    table.set_value((2, 0), 'world')
    document.body.append(Table('Second'))
    document.save(file_path)


def _create_styled_ods(file_path: Path) -> None:
    """Create one styled `.ods` file for checker tests."""
    document = Document('spreadsheet')
    document.body.clear()
    style_name = _yellow_ods_style(document)
    table = Table('Sheet1')
    document.body.append(table)
    styled_hello = Cell('hello')
    styled_hello.style = style_name
    table.set_cell((0, 0), styled_hello, clone=False)
    styled_wonderful = Cell('wonderful')
    styled_wonderful.style = style_name
    table.set_cell((1, 0), styled_wonderful, clone=False)
    styled_world = Cell('world')
    styled_world.style = style_name
    table.set_cell((2, 0), styled_world, clone=False)
    table.set_cell((0, 1), Cell('hello'), clone=False)
    table.set_cell((1, 1), Cell('world'), clone=False)
    document.save(file_path)


def _border_helper(border_style: TableBorderStyle) -> BorderHelper:
    """Return one border helper with border-writing capability enabled."""
    return BorderHelper(border_style,
                        Capabilities(can_write_borders=CAP_NEEDED))


_XLSX_BORDER_CODES: dict[BorderWeight, int] = {
    BorderWeight.THIN: 1,
    BorderWeight.THICK: 2
}


def _xlsx_border_code(weight: BorderWeight) -> Optional[int]:
    """Return one XlsxWriter border code for one normalized weight."""
    return _XLSX_BORDER_CODES.get(weight)


def _create_bordered_xlsx(file_path: Path) -> None:
    """Create one `.xlsx` file with a bordered 2 x 2 table."""
    workbook = Workbook(file_path)
    worksheet = workbook.add_worksheet('Sheet1')
    border_helper = _border_helper(TableBorderStyle.OUTER_THICK_INNER_THIN)
    values = [['hello', 'world'], ['later', 'done']]
    for row_index, row in enumerate(values):
        for col_index, value in enumerate(row):
            borders = border_helper.cell_border(row_index, col_index, 2, 2)
            format_dict: dict[str, object] = {}
            for side_name, weight in [('left', borders.left),
                                      ('right', borders.right),
                                      ('top', borders.top),
                                      ('bottom', borders.bottom)]:
                border_code = _xlsx_border_code(weight)
                if border_code is not None:
                    format_dict[side_name] = border_code
            cell_format = workbook.add_format(format_dict)
            worksheet.write(row_index, col_index, value, cell_format)
    workbook.close()


def _ods_border_text(weight: BorderWeight) -> Optional[str]:
    """Return one ODS border property value for one normalized weight."""
    if weight == BorderWeight.NONE:
        return None
    if weight == BorderWeight.THIN:
        return '0.75pt solid #000000'
    return '1.75pt solid #000000'


def _ods_style_name(document: Document,
                    borders: CellBorder,
                    style_names: dict[CellBorder, str]) -> str:
    """Return one cached ODS cell style name for one cell border."""
    cached = style_names.get(borders)
    if cached is not None:
        return cached
    style_name = f'ce_border_style_{len(style_names)}'
    style = Style('table-cell', name=style_name)
    table_props: dict[str, Any] = {}
    for property_name, weight in [('fo:border-top', borders.top),
                                  ('fo:border-right', borders.right),
                                  ('fo:border-bottom', borders.bottom),
                                  ('fo:border-left', borders.left)]:
        border_text = _ods_border_text(weight)
        if border_text is not None:
            table_props[property_name] = border_text
    if table_props:
        style.set_properties(table_props, area='table-cell')
    document.insert_style(style, automatic=True)
    style_names[borders] = style_name
    return style_name


def _create_bordered_ods(file_path: Path) -> None:
    """Create one `.ods` file with a bordered 2 x 2 table."""
    document = Document('spreadsheet')
    document.body.clear()
    table = Table('Sheet1')
    document.body.append(table)
    border_helper = _border_helper(TableBorderStyle.OUTER_THICK_INNER_THIN)
    style_names: dict[CellBorder, str] = {}
    values = [['hello', 'world'], ['later', 'done']]
    for row_index, row in enumerate(values):
        for col_index, value in enumerate(row):
            cell = Cell(value)
            cell.style = _ods_style_name(
                document,
                border_helper.cell_border(row_index, col_index, 2, 2),
                style_names
            )
            table.set_cell((col_index, row_index), cell, clone=False)
    document.save(file_path)


def _create_invalid_xlsx_with_custom_font(file_path: Path) -> None:
    """Create one `.xlsx` file with a validator-rejected font definition."""
    workbook = OpenPyXLWorkbook()
    worksheet = workbook.active
    assert isinstance(worksheet, Worksheet)
    worksheet['A1'] = 1
    worksheet['A1'].font = Font(bold=True, size=14)
    workbook.save(file_path)
    workbook.close()


def test_expected_cell_style_defaults_to_dont_care() -> None:
    """Test ExpectedCellStyle defaults every style aspect to unchecked."""
    style = ExpectedCellStyle()
    assert style.bold is None
    assert style.italic is None
    assert style.background_color is None


def test_spreadsheet_output_path_appends_suffix_once() -> None:
    """Test spreadsheet_output_path appends one normalized suffix."""
    path = _spreadsheet_output_path(Path('tmp') / 'example', 'xlsx')
    assert path == Path('tmp') / 'example.xlsx'


def test_check_spreadsheet_syntax_accepts_valid_xlsx(tmp_path: Path) -> None:
    """Valid `.xlsx` files pass syntax validation."""
    file_path = tmp_path / 'example.xlsx'
    _create_basic_xlsx(file_path)
    check_spreadsheet_syntax(file_path)


def test_check_spreadsheet_content_matches_subsequence_in_xlsx(
        tmp_path: Path) -> None:
    """Content checks accept subsequence row fragments in `.xlsx` files."""
    file_path = tmp_path / 'example.xlsx'
    _create_basic_xlsx(file_path)
    check_spreadsheet_content(
        file_path,
        [
            SheetContentExpectation(
                sheet_name='Sheet1',
                row_fragments=[['hello', 'world']]
            ),
            SheetContentExpectation(
                sheet_name='Second',
                row_fragments=[]
            )
        ]
    )


def test_check_spreadsheet_content_matches_typed_values_in_xlsx(
        tmp_path: Path) -> None:
    """Content checks can match numeric and boolean fragments in `.xlsx`."""
    file_path = tmp_path / 'example.xlsx'
    workbook = Workbook(file_path)
    worksheet = workbook.add_worksheet('Sheet1')
    worksheet.write_row(0, 0, ['hello', 'wonderful', 'world'])
    worksheet.write_string(1, 0, 'later')
    worksheet.write_row(2, 0, [3.14159, True, 'done'])
    workbook.add_worksheet('Second')
    workbook.close()
    check_spreadsheet_content(
        file_path,
        [
            SheetContentExpectation(
                sheet_name='Sheet1',
                row_fragments=[['hello', 'world'], [3.14159, True]]
            ),
            SheetContentExpectation(
                sheet_name='Second',
                row_fragments=[]
            )
        ]
    )


def test_check_spreadsheet_content_matches_subsequence_in_ods(
        tmp_path: Path) -> None:
    """Content checks accept subsequence row fragments in `.ods` files."""
    file_path = tmp_path / 'example.ods'
    _create_basic_ods(file_path)
    check_spreadsheet_content(
        file_path,
        [
            SheetContentExpectation(
                sheet_name='Sheet1',
                row_fragments=[['hello', 'world']]
            ),
            SheetContentExpectation(
                sheet_name='Second',
                row_fragments=[]
            )
        ]
    )


def test_check_spreadsheet_syntax_accepts_expected_error_fragments(
        tmp_path: Path) -> None:
    """Expected validator errors can be allowed explicitly."""
    file_path = tmp_path / 'invalid.xlsx'
    _create_invalid_xlsx_with_custom_font(file_path)
    check_spreadsheet_syntax(file_path,
                             expected_errors=['Unexpected element'])


def test_check_spreadsheet_syntax_rejects_missing_expected_errors(
        tmp_path: Path) -> None:
    """Expected-error matching fails when a fragment is not reported."""
    file_path = tmp_path / 'invalid.xlsx'
    _create_invalid_xlsx_with_custom_font(file_path)
    with pytest.raises(AssertionError, match='Missing expected errors'):
        check_spreadsheet_syntax(file_path,
                                 expected_errors=['not reported'])


def test_check_spreadsheet_content_checks_sheet_order(
        tmp_path: Path) -> None:
    """Content checks fail when sheet names are in the wrong order."""
    file_path = tmp_path / 'example.xlsx'
    _create_basic_xlsx(file_path)
    with pytest.raises(AssertionError, match='Unexpected sheet names'):
        check_spreadsheet_content(
            file_path,
            [
                SheetContentExpectation(
                    sheet_name='Second',
                    row_fragments=[]
                ),
                SheetContentExpectation(
                    sheet_name='Sheet1',
                    row_fragments=[['hello', 'world']]
                )
            ]
        )


def test_check_spreadsheet_styles_uses_default_anchor_offset_in_xlsx(
        tmp_path: Path) -> None:
    """Style checks default to the anchor cell in `.xlsx` files."""
    file_path = tmp_path / 'example.xlsx'
    _create_styled_xlsx(file_path)
    check_spreadsheet_styles(
        file_path,
        [
            AnchoredStyleExpectation(
                sheet_name='Sheet1',
                anchor_row_fragment=['hello', 'world'],
                relative_expectations=[
                    RelativeStyleExpectation(
                        expected_style=ExpectedCellStyle(
                            bold=True,
                            italic=True,
                            background_color=Color.YELLOW
                        )
                    )
                ]
            )
        ]
    )


def test_check_spreadsheet_styles_checks_rectangular_area_in_xlsx(
        tmp_path: Path) -> None:
    """Style checks can verify one rectangular area in `.xlsx` files."""
    file_path = tmp_path / 'example.xlsx'
    _create_styled_xlsx(file_path)
    check_spreadsheet_styles(
        file_path,
        [
            AnchoredStyleExpectation(
                sheet_name='Sheet1',
                anchor_row_fragment=['hello', 'world'],
                relative_expectations=[
                    RelativeStyleExpectation(
                        expected_style=ExpectedCellStyle(
                            bold=True,
                            italic=True,
                            background_color=Color.YELLOW
                        ),
                        number_of_columns=3
                    )
                ]
            )
        ]
    )


def test_check_spreadsheet_styles_uses_default_anchor_offset_in_ods(
        tmp_path: Path) -> None:
    """Style checks default to the anchor cell in `.ods` files."""
    file_path = tmp_path / 'example.ods'
    _create_styled_ods(file_path)
    check_spreadsheet_styles(
        file_path,
        [
            AnchoredStyleExpectation(
                sheet_name='Sheet1',
                anchor_row_fragment=['hello', 'world'],
                relative_expectations=[
                    RelativeStyleExpectation(
                        expected_style=ExpectedCellStyle(
                            bold=True,
                            italic=True,
                            background_color=Color.YELLOW
                        )
                    )
                ]
            )
        ]
    )


def test_check_spreadsheet_styles_checks_rectangular_area_in_ods(
        tmp_path: Path) -> None:
    """Style checks can verify one rectangular area in `.ods` files."""
    file_path = tmp_path / 'example.ods'
    _create_styled_ods(file_path)
    check_spreadsheet_styles(
        file_path,
        [
            AnchoredStyleExpectation(
                sheet_name='Sheet1',
                anchor_row_fragment=['hello', 'world'],
                relative_expectations=[
                    RelativeStyleExpectation(
                        expected_style=ExpectedCellStyle(
                            bold=True,
                            italic=True,
                            background_color=Color.YELLOW
                        ),
                        number_of_columns=3
                    )
                ]
            )
        ]
    )


def test_check_spreadsheet_borders_checks_table_style_in_xlsx(
        tmp_path: Path) -> None:
    """Border checks can verify one table style in `.xlsx` files."""
    file_path = tmp_path / 'example.xlsx'
    _create_bordered_xlsx(file_path)
    check_spreadsheet_borders(
        file_path,
        [
            AnchoredBorderExpectation(
                sheet_name='Sheet1',
                anchor_row_fragment=['hello', 'world'],
                relative_expectations=[
                    RelativeBorderExpectation(
                        border_style=TableBorderStyle.OUTER_THICK_INNER_THIN,
                        number_of_rows=2,
                        number_of_columns=2
                    )
                ]
            )
        ]
    )


def test_check_spreadsheet_borders_checks_table_style_in_ods(
        tmp_path: Path) -> None:
    """Border checks can verify one table style in `.ods` files."""
    file_path = tmp_path / 'example.ods'
    _create_bordered_ods(file_path)
    check_spreadsheet_borders(
        file_path,
        [
            AnchoredBorderExpectation(
                sheet_name='Sheet1',
                anchor_row_fragment=['hello', 'world'],
                relative_expectations=[
                    RelativeBorderExpectation(
                        border_style=TableBorderStyle.OUTER_THICK_INNER_THIN,
                        number_of_rows=2,
                        number_of_columns=2
                    )
                ]
            )
        ]
    )


@pytest.mark.parametrize(
    ('number_of_rows', 'number_of_columns', 'expected_message'),
    [
        (0, 1, 'number_of_rows'),
        (-1, 1, 'number_of_rows'),
        (1, 0, 'number_of_columns'),
        (1, -1, 'number_of_columns')
    ]
)
def test_check_spreadsheet_styles_rejects_non_positive_area_sizes(
        tmp_path: Path,
        number_of_rows: int,
        number_of_columns: int,
        expected_message: str) -> None:
    """Style checks reject non-positive rectangular area sizes."""
    file_path = tmp_path / 'example.xlsx'
    _create_styled_xlsx(file_path)
    with pytest.raises(ValueError, match=expected_message):
        check_spreadsheet_styles(
            file_path,
            [
                AnchoredStyleExpectation(
                    sheet_name='Sheet1',
                    anchor_row_fragment=['hello', 'world'],
                    relative_expectations=[
                        RelativeStyleExpectation(
                            expected_style=ExpectedCellStyle(bold=True),
                            number_of_rows=number_of_rows,
                            number_of_columns=number_of_columns
                        )
                    ]
                )
            ]
        )


def test_check_spreadsheet_styles_reports_all_area_mismatches(
        tmp_path: Path) -> None:
    """Style mismatch reports include all mismatches in one area."""
    file_path = tmp_path / 'example.xlsx'
    _create_basic_xlsx(file_path)
    with pytest.raises(AssertionError) as error_info:
        check_spreadsheet_styles(
            file_path,
            [
                AnchoredStyleExpectation(
                    sheet_name='Sheet1',
                    anchor_row_fragment=['hello', 'world'],
                    relative_expectations=[
                        RelativeStyleExpectation(
                            expected_style=ExpectedCellStyle(bold=True),
                            number_of_columns=2
                        )
                    ]
                )
            ]
        )
    assert 'Style mismatches in area' in str(error_info.value)
    assert 'row 1, column 1' in str(error_info.value)
    assert 'row 1, column 2' in str(error_info.value)
    assert str(error_info.value).count('Unexpected bold value') == 2


def test_check_spreadsheet_borders_reports_all_edge_mismatches(
        tmp_path: Path) -> None:
    """Border mismatch reports include the failing edge details."""
    file_path = tmp_path / 'example.xlsx'
    _create_basic_xlsx(file_path)
    with pytest.raises(AssertionError) as error_info:
        check_spreadsheet_borders(
            file_path,
            [
                AnchoredBorderExpectation(
                    sheet_name='Sheet1',
                    anchor_row_fragment=['hello', 'world'],
                    relative_expectations=[
                        RelativeBorderExpectation(
                            border_style=TableBorderStyle.OUTER_THICK,
                            number_of_rows=1,
                            number_of_columns=2
                        )
                    ]
                )
            ]
        )
    assert 'Border mismatches in area' in str(error_info.value)
    assert 'Unexpected top border' in str(error_info.value)
    assert 'row 1, column 1' in str(error_info.value)


def test_check_spreadsheet_styles_treats_missing_cells_as_unformatted(
        tmp_path: Path) -> None:
    """Style checks treat missing cells outside content as unformatted."""
    file_path = tmp_path / 'example.xlsx'
    _create_basic_xlsx(file_path)
    check_spreadsheet_styles(
        file_path,
        [
            AnchoredStyleExpectation(
                sheet_name='Sheet1',
                anchor_row_fragment=['hello', 'world'],
                relative_expectations=[
                    RelativeStyleExpectation(
                        expected_style=ExpectedCellStyle(
                            bold=False,
                            italic=False,
                            background_color=Color.NONE
                        ),
                        number_of_columns=5
                    )
                ]
            )
        ]
    )


def test_check_spreadsheet_file_runs_full_xlsx_validation(
        tmp_path: Path) -> None:
    """The combined checker validates syntax, content and style for `.xlsx`."""
    file_path = tmp_path / 'example.xlsx'
    _create_styled_xlsx(file_path)
    check_spreadsheet_file(
        file_path,
        [
            SheetContentExpectation(
                sheet_name='Sheet1',
                row_fragments=[['hello', 'world']]
            ),
            SheetContentExpectation(
                sheet_name='Second',
                row_fragments=[]
            )
        ],
        [
            AnchoredStyleExpectation(
                sheet_name='Sheet1',
                anchor_row_fragment=['hello', 'world'],
                relative_expectations=[
                    RelativeStyleExpectation(
                        expected_style=ExpectedCellStyle(
                            bold=True,
                            italic=True,
                            background_color=Color.YELLOW
                        ),
                        number_of_columns=3
                    )
                ]
            )
        ]
    )


def test_check_spreadsheet_file_runs_full_xlsx_validation_with_borders(
        tmp_path: Path) -> None:
    """The combined checker also validates border expectations."""
    file_path = tmp_path / 'example.xlsx'
    _create_bordered_xlsx(file_path)
    check_spreadsheet_file(
        file_path,
        [
            SheetContentExpectation(
                sheet_name='Sheet1',
                row_fragments=[['hello', 'world'], ['later', 'done']]
            )
        ],
        border_expectations=[
            AnchoredBorderExpectation(
                sheet_name='Sheet1',
                anchor_row_fragment=['hello', 'world'],
                relative_expectations=[
                    RelativeBorderExpectation(
                        border_style=TableBorderStyle.OUTER_THICK_INNER_THIN,
                        number_of_rows=2,
                        number_of_columns=2
                    )
                ]
            )
        ]
    )
