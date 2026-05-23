#! /usr/local/bin/python3
"""Tests for the tableio_excel_pylightxl module."""

# Copyright (c) 2026 Tom Björkholm
# MIT License

from datetime import datetime
from pathlib import Path
from tempfile import TemporaryDirectory
import xml.etree.ElementTree as ET
from zipfile import ZipFile
from pylightxl import Database  # type: ignore[import-untyped]
import pytest
from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from openxml_audit import OpenXmlValidator  # type: ignore[import-untyped]
from pytest import CaptureFixture
import tableio.tableio_excel_pylightxl as tableio_excel_pylightxl_module
from tableio.capability import CAP_IGNORED, CAP_IMPLEMENTED
from tableio.color import Color
from tableio.factory import create_tableio
from tableio.tableio import Box, FileAccess
from tableio.tableio_excel_openpyxl import TableIOExcelOpenPyXL
from tableio.tableio_excel_pylightxl import TableIOExcelPylightxl, \
    _datetime_to_excel_number as datetime_to_excel_number, \
    _load_named_ranges as load_named_ranges, \
    _number_from_cell_text as number_from_cell_text, \
    _sheet_data_from_xml as sheet_data_from_xml, \
    _sheet_xml_targets as sheet_xml_targets, \
    _style_index_for_code as style_index_for_code, \
    _worksheet_id_attr as worksheet_id_attr, \
    _XML_NS as XML_NS
from tableio.value_type import Fmt, FmtDictRow, ValueFmt
from .check_capsys import check_capsys
from .excel_test_file_helper import create_formula_workbook, \
    create_update_workbook, inspect_find_and_write_cells_workbook, \
    inspect_updated_workbook
from .spreadsheet_test_helper import \
    run_box_partial_overwrite, \
    run_find_value_and_write_cells, \
    run_multi_sheet_heading_state_is_per_sheet, \
    run_multi_sheet_read_only_create_raises, \
    run_multi_sheet_read_positions_are_per_sheet, \
    run_multi_sheet_update_uses_selected_sheet_write_position, \
    run_multi_sheet_write_positions_are_per_sheet, \
    run_open_rejects_second_open, \
    run_read_formula_cached, \
    run_read_formula_no_cache, \
    run_round_trip_dictdata_in_box, \
    run_sequential_list_reads, \
    run_select_missing_sheet_without_create_raises_key_error, \
    run_update_default_write_starts_after_last_used_row


def _write_zip_members(file_name: Path, members: dict[str, bytes]) -> None:
    """Write one ZIP file from named byte members."""
    with ZipFile(file_name, 'w') as zip_file:
        for member_name, data in members.items():
            zip_file.writestr(member_name, data)


def _rewrite_workbook_members() -> dict[str, bytes]:
    """Return a minimal workbook archive used by XML rewrite tests."""
    return {
        '[Content_Types].xml': (
            '<?xml version="1.0" encoding="UTF-8"?>'
            '<Types '
            'xmlns="http://schemas.openxmlformats.org/package/2006/'
            'content-types">'
            '<Override PartName="/xl/workbook.xml" '
            'ContentType="application/vnd.openxmlformats-officedocument.'
            'spreadsheetml.sheet.main+xml"/>'
            '<Override PartName="/xl/styles.xml" ContentType="text/plain"/>'
            '</Types>'
        ).encode('utf-8'),
        'xl/_rels/workbook.xml.rels': (
            '<?xml version="1.0" encoding="UTF-8"?>'
            '<Relationships '
            'xmlns="http://schemas.openxmlformats.org/package/2006/'
            'relationships">'
            '<Relationship Id="rId1" '
            'Type="http://schemas.openxmlformats.org/officeDocument/2006/'
            'relationships/worksheet" '
            'Target="worksheets/sheet1.xml"/>'
            '<Relationship Id="rId9" '
            'Type="http://schemas.openxmlformats.org/officeDocument/2006/'
            'relationships/styles" '
            'Target="oldstyles.xml"/>'
            '</Relationships>'
        ).encode('utf-8'),
        'xl/workbook.xml': (
            '<?xml version="1.0" encoding="UTF-8"?>'
            '<workbook '
            'xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/'
            'main">'
            '<sheets><sheet name="Sheet1" sheetId="1" id="rId1"/>'
            '</sheets>'
            '</workbook>'
        ).encode('utf-8'),
        'xl/worksheets/sheet1.xml': (
            '<?xml version="1.0" encoding="UTF-8"?>'
            '<worksheet '
            'xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/'
            'main">'
            '<sheetData><row r="1">'
            '<c r="A1"><v>None</v></c>'
            '<c r="A2"><v>False</v></c>'
            '<c><v>7</v></c>'
            '<c r="A3"><v>9</v></c>'
            '</row></sheetData>'
            '</worksheet>'
        ).encode('utf-8'),
        'xl/styles.xml': b'old styles'
    }


class _InspectableTableIOExcelPylightxl(TableIOExcelPylightxl):
    """Expose protected pylightxl helpers needed by the tests."""

    @classmethod
    def parse_typed_value(cls, value: object, style_code: str) -> object:
        """Expose the typed-cell conversion helper for tests."""
        return cls._parse_typed_cell_value(value, style_code)

    @staticmethod
    def invalid_placeholder_cell(cell: ET.Element) -> bool:
        """Expose the placeholder-cell helper for tests."""
        return TableIOExcelPylightxl._invalid_placeholder_cell(cell)

    @staticmethod
    def normalize_written_bool_cell(cell: ET.Element) -> bool:
        """Expose the bool-normalization helper for tests."""
        return TableIOExcelPylightxl._normalize_written_bool_cell(cell)

    def run_last_used_column(self) -> int:
        """Expose the last-used-column helper for tests."""
        return self._last_used_column(self._write_sheet())

    def run_delete_filtered_range(self, name: str) -> None:
        """Expose filtered-range deletion for tests."""
        self._delete_filtered_range(name)

    def run_add_filtered_range(self, bounds: tuple[int, int, int, int],
                               name: str) -> None:
        """Expose filtered-range creation for tests."""
        self._add_filtered_range(bounds, name)

    def run_current_sheet_name(self) -> str:
        """Expose current-sheet lookup for tests."""
        return self._current_sheet_name()

    def run_close_hook(self) -> None:
        """Expose the close hook for tests."""
        self._close()

    def run_write_file_suffix(self) -> None:
        """Expose workbook finalization for tests."""
        self._write_file_suffix()

    def set_sheet_style_codes(self,
                              style_codes: dict[str, dict[str, str]]) -> None:
        """Expose sheet style metadata mutation for tests."""
        self._sheet_style_codes = style_codes

    def run_entry_style_codes(
            self, entry_name: str,
            sheet_targets: dict[str, tuple[int, str]]) -> dict[str, str]:
        """Expose sheet-entry style lookup for tests."""
        return self._entry_style_codes(entry_name, sheet_targets)

    def run_worksheet_xml_for_output(
            self, entry_name: str, data: bytes,
            sheet_targets: dict[str, tuple[int, str]]) -> bytes:
        """Expose worksheet XML cleanup for tests."""
        return self._worksheet_xml_for_output(entry_name, data, sheet_targets)

    def run_rewrite_workbook_xml(self, file_name: Path) -> None:
        """Expose workbook XML rewrite for tests."""
        self._rewrite_workbook_xml(file_name)


def test_pylightxl_capabilities(capsys: CaptureFixture[str]) -> None:
    """The backend reports the honest pylightxl feature set."""
    capabilities = TableIOExcelPylightxl.get_capabilities()
    assert capabilities.can_read == CAP_IMPLEMENTED
    assert capabilities.can_write == CAP_IMPLEMENTED
    assert capabilities.can_fmt_row == CAP_IGNORED
    assert capabilities.can_fmt_value == CAP_IGNORED
    assert capabilities.filtered_data_range == CAP_IGNORED
    assert capabilities.can_write_box == CAP_IMPLEMENTED
    assert capabilities.can_read_box == CAP_IMPLEMENTED
    assert capabilities.can_write_highlight == CAP_IGNORED
    assert capabilities.multi_sheet == CAP_IMPLEMENTED
    assert capabilities.can_find_value_position == CAP_IMPLEMENTED
    assert capabilities.can_write_borders == CAP_IGNORED
    check_capsys(capsys)


def test_excel_pylightxl_get_description_uses_lower_priority(
        capsys: CaptureFixture[str]) -> None:
    """The backend stays below OpenPyXL in default factory preference."""
    description = TableIOExcelPylightxl.get_description()
    assert description.format_name == 'Excel'
    assert description.implementation == 'pylightxl'
    assert description.priority == 8
    assert description.priority < TableIOExcelOpenPyXL.get_description(
    ).priority
    check_capsys(capsys)


def test_excel_factory_can_create_pylightxl_explicitly(
        capsys: CaptureFixture[str]) -> None:
    """The factory can create the backend when requested explicitly."""
    with TemporaryDirectory() as temp_dir:
        table_io = create_tableio('Excel', Path(temp_dir) / 'explicit',
                                  FileAccess.CREATE,
                                  implementation='pylightxl')
        assert isinstance(table_io, TableIOExcelPylightxl)
    check_capsys(capsys)


def test_excel_pylightxl_round_trip_sequential_list_reads(
        capsys: CaptureFixture[str]) -> None:
    """Two list sections can be written and then read back sequentially."""
    run_sequential_list_reads(TableIOExcelPylightxl, capsys)


def test_excel_pylightxl_round_trip_dictdata_in_box(
        capsys: CaptureFixture[str]) -> None:
    """Dict data can be written into and read back from a box."""
    run_round_trip_dictdata_in_box(TableIOExcelPylightxl, capsys)


def test_excel_pylightxl_multi_sheet_write_positions_are_per_sheet(
        capsys: CaptureFixture[str]) -> None:
    """Sequential writes keep an independent default position per sheet."""
    run_multi_sheet_write_positions_are_per_sheet(TableIOExcelPylightxl,
                                                  capsys)


def test_excel_pylightxl_written_workbook_is_validator_clean() -> None:
    """Plain `.xlsx` output includes the required validator metadata."""
    with TemporaryDirectory() as temp_dir:
        file_path = Path(temp_dir) / 'validator_clean.xlsx'
        with TableIOExcelPylightxl(file_path, FileAccess.CREATE) as table_io:
            table_io.write_table_listdata([['left', 'right']])
        result = OpenXmlValidator().validate(file_path)
        assert result.is_valid


def test_excel_pylightxl_multi_sheet_read_positions_are_per_sheet(
        capsys: CaptureFixture[str]) -> None:
    """Sequential reads resume independently when switching sheets."""
    run_multi_sheet_read_positions_are_per_sheet(TableIOExcelPylightxl, capsys)


def test_excel_pylightxl_multi_sheet_heading_state_is_per_sheet(
        capsys: CaptureFixture[str]) -> None:
    """Each sheet tracks whether a default heading level was used before."""
    run_multi_sheet_heading_state_is_per_sheet(TableIOExcelPylightxl, capsys)


def test_excel_pylightxl_multi_sheet_read_only_create_raises(
        capsys: CaptureFixture[str]) -> None:
    """READ mode can select an existing sheet but cannot create one."""
    run_multi_sheet_read_only_create_raises(TableIOExcelPylightxl, capsys)


def test_excel_pylightxl_update_default_write_starts_after_last_used_row(
        capsys: CaptureFixture[str]) -> None:
    """UPDATE mode appends after the used area with a blank row separator."""
    run_update_default_write_starts_after_last_used_row(
        TableIOExcelPylightxl, '.xlsx', create_update_workbook,
        inspect_updated_workbook, capsys)


def test_excel_pylightxl_multi_sheet_update_uses_selected_sheet_write_position(
        capsys: CaptureFixture[str]) -> None:
    """UPDATE mode appends using the selected sheet's used area."""
    run_multi_sheet_update_uses_selected_sheet_write_position(
        TableIOExcelPylightxl, capsys)


def test_excel_pylightxl_find_value_and_write_cells(
        capsys: CaptureFixture[str]) -> None:
    """Found cell ranges can be read and updated without moving cursors."""
    run_find_value_and_write_cells(TableIOExcelPylightxl, '.xlsx',
                                   inspect_find_and_write_cells_workbook,
                                   capsys)


def test_excel_pylightxl_boxed_table_partial_overwrite_raises(
        capsys: CaptureFixture[str]) -> None:
    """Boxed table writes reject overlaps that leave part of a table behind."""
    run_box_partial_overwrite(TableIOExcelPylightxl, capsys)


def test_excel_pylightxl_read_formula_uses_cached_value(
        capsys: CaptureFixture[str]) -> None:
    """A formula cell is read as its cached value."""
    run_read_formula_cached(TableIOExcelPylightxl, '.xlsx',
                            create_formula_workbook, 3, capsys)


def test_excel_pylightxl_read_formula_without_cached_value_returns_none(
        capsys: CaptureFixture[str]) -> None:
    """A formula without a cached result is read as None."""
    run_read_formula_no_cache(TableIOExcelPylightxl, '.xlsx',
                              create_formula_workbook, capsys)


def test_excel_pylightxl_open_rejects_second_open(
        capsys: CaptureFixture[str]) -> None:
    """Opening the same Excel object twice raises RuntimeError."""
    run_open_rejects_second_open(TableIOExcelPylightxl, capsys)


def test_excel_pylightxl_select_missing_sheet_without_create_raises(
        capsys: CaptureFixture[str]) -> None:
    """Selecting a missing sheet without create=True raises KeyError."""
    run_select_missing_sheet_without_create_raises_key_error(
        TableIOExcelPylightxl, capsys)


def test_excel_pylightxl_ignored_format_and_filter_requests_keep_data(
        capsys: CaptureFixture[str]) -> None:
    """Formatting and filtered-range requests are ignored but data is kept."""
    with TemporaryDirectory() as temp_dir:
        file_name = Path(temp_dir) / 'ignored_features'
        with TableIOExcelPylightxl(file_name, FileAccess.CREATE) as table_io:
            table_io.write_table_listdata([
                [ValueFmt(value='name', fmt=Fmt(bold=True)),
                 ValueFmt(value='active', fmt=Fmt(highlight=Color.YELLOW))],
                [ValueFmt(value='Alice',
                          fmt=Fmt(italic=True, highlight=Color.GREEN)),
                 ValueFmt(value=True,
                          fmt=Fmt(italic=True, highlight=Color.GREEN))]
            ], filtered_data_range=True)
            table_io.write_table_fmtdictdata(
                data=[FmtDictRow(values={'name': 'Bob', 'active': False},
                                 fmt=Fmt(bold=True))],
                column_order=['name', 'active'],
                first_row_format=Fmt(bold=True), filtered_data_range=True,
                box=Box(top=4, left=0, bottom=6, right=2))
        workbook = load_workbook(Path(temp_dir) / 'ignored_features.xlsx')
        worksheet = workbook.active
        assert isinstance(worksheet, Worksheet)
        assert worksheet['A1'].value == 'name'
        assert worksheet['B2'].value is True
        assert worksheet['A5'].value == 'name'
        assert worksheet['B6'].value is False
        assert not worksheet.tables
        assert worksheet['A1'].font.bold is False
        assert worksheet['A2'].font.italic is False
        workbook.close()
    check_capsys(capsys)


def test_excel_pylightxl_reads_openpyxl_datetime_cells(
        capsys: CaptureFixture[str]) -> None:
    """The read workaround handles ordinary OpenPyXL datetime cells."""
    with TemporaryDirectory() as temp_dir:
        file_name = Path(temp_dir) / 'openpyxl_datetime.xlsx'
        workbook = Workbook()
        worksheet = workbook.active
        assert isinstance(worksheet, Worksheet)
        when = datetime(2026, 3, 24, 14, 30, 0)
        worksheet['A1'] = when
        worksheet['B1'] = 'x'
        workbook.save(file_name)
        workbook.close()
        with TableIOExcelPylightxl(Path(temp_dir) / 'openpyxl_datetime',
                                   FileAccess.READ) as table_io:
            result = table_io.read_table_listdata()
        assert result.data == [[when, 'x']]
    check_capsys(capsys)


def test_excel_pylightxl_writes_datetime_readably_and_reads_it_back(
        capsys: CaptureFixture[str]) -> None:
    """Written datetime cells use ISO display format and round-trip here."""
    with TemporaryDirectory() as temp_dir:
        file_name = Path(temp_dir) / 'datetime_roundtrip'
        when = datetime(2026, 3, 24, 14, 30, 0)
        with TableIOExcelPylightxl(file_name, FileAccess.CREATE) as table_io:
            table_io.write_table_listdata([['when'], [when]])
        workbook = load_workbook(Path(temp_dir) / 'datetime_roundtrip.xlsx')
        worksheet = workbook.active
        assert isinstance(worksheet, Worksheet)
        assert worksheet['A2'].value == when
        assert worksheet['A2'].number_format == 'yyyy-mm-dd hh:mm:ss'
        workbook.close()
        with TableIOExcelPylightxl(Path(temp_dir) / 'datetime_roundtrip',
                                   FileAccess.READ) as table_io:
            result = table_io.read_table_listdata()
        assert result.data == [['when'], [when]]
    check_capsys(capsys)


def test_excel_pylightxl_update_keeps_datetime_cells_readable(
        capsys: CaptureFixture[str]) -> None:
    """UPDATE mode preserves readable datetime data while appending rows."""
    with TemporaryDirectory() as temp_dir:
        file_name = Path(temp_dir) / 'update_datetime.xlsx'
        workbook = Workbook()
        worksheet = workbook.active
        assert isinstance(worksheet, Worksheet)
        when = datetime(2026, 3, 24, 14, 30, 0)
        worksheet['A1'] = when
        worksheet['B1'] = 'start'
        workbook.save(file_name)
        workbook.close()
        with TableIOExcelPylightxl(Path(temp_dir) / 'update_datetime',
                                   FileAccess.UPDATE) as table_io:
            table_io.write_table_listdata([['new', 'row']])
        workbook = load_workbook(file_name)
        worksheet = workbook.active
        assert isinstance(worksheet, Worksheet)
        assert worksheet['A1'].value == when
        assert worksheet['B1'].value == 'start'
        assert worksheet['A3'].value == 'new'
        workbook.close()
        with TableIOExcelPylightxl(Path(temp_dir) / 'update_datetime',
                                   FileAccess.READ) as table_io:
            first_result = table_io.read_table_listdata()
            second_result = table_io.read_table_listdata()
        assert first_result.data == [[when, 'start']]
        assert second_result.data == [['new', 'row']]
    check_capsys(capsys)


def test_excel_pylightxl_rewrites_worksheet_xml_with_excel_prefixes(
        capsys: CaptureFixture[str]) -> None:
    """Rewritten worksheet XML keeps namespace prefixes Excel accepts."""
    with TemporaryDirectory() as temp_dir:
        file_name = Path(temp_dir) / 'xml_prefixes'
        with TableIOExcelPylightxl(file_name, FileAccess.CREATE) as table_io:
            table_io.write_table_listdata([['value'], [True]])
        with ZipFile(Path(temp_dir) / 'xml_prefixes.xlsx') as zip_file:
            worksheet_xml = zip_file.read(
                'xl/worksheets/sheet1.xml').decode('utf-8')
        assert 'mc:Ignorable="x14ac xr"' in worksheet_xml
        assert 'xmlns:mc="' in worksheet_xml
        assert 'xmlns:x14ac="' in worksheet_xml
        assert 'xmlns:xr="' in worksheet_xml
    check_capsys(capsys)


def test_excel_pylightxl_xml_target_helpers_cover_id_fallbacks(
        capsys: CaptureFixture[str]) -> None:
    """Worksheet id and target helpers accept multiple workbook layouts."""
    rels_namespace = XML_NS['rels']
    assert worksheet_id_attr(
        ET.Element('sheet', {'{%s}id' % rels_namespace: 'pkgRel'})) == \
        'pkgRel'
    assert worksheet_id_attr(
        ET.Element('sheet', {'{urn:test}id': 'otherRel'})) == 'otherRel'
    assert worksheet_id_attr(
        ET.Element('sheet', {'id': 'plainRel'})) == 'plainRel'
    with TemporaryDirectory() as temp_dir:
        file_name = Path(temp_dir) / 'sheet_targets.xlsx'
        workbook_xml = (
            '<?xml version="1.0" encoding="UTF-8"?>'
            '<workbook '
            'xmlns="http://schemas.openxmlformats.org/'
            'spreadsheetml/2006/main">'
            '<sheets>'
            '<sheet name="First" sheetId="5" id="rId1"/>'
            '<sheet name="NoTarget" sheetId="6" id="rId2"/>'
            '<sheet name="MissingId" id="rId3"/>'
            '</sheets>'
            '</workbook>'
        ).encode('utf-8')
        rels_xml = (
            '<?xml version="1.0" encoding="UTF-8"?>'
            '<Relationships '
            'xmlns="http://schemas.openxmlformats.org/package/2006/'
            'relationships">'
            '<Relationship Id="rId1" Target="/xl/worksheets/sheet1.xml"/>'
            '<Relationship Id="rId2"/>'
            '<Relationship Target="worksheets/sheet3.xml"/>'
            '</Relationships>'
        ).encode('utf-8')
        _write_zip_members(file_name, {
            'xl/workbook.xml': workbook_xml,
            'xl/_rels/workbook.xml.rels': rels_xml
        })
        assert sheet_xml_targets(str(file_name)) == {
            'First': (5, 'worksheets/sheet1.xml')
        }
    check_capsys(capsys)


def test_excel_pylightxl_sheet_data_parsing_handles_edge_cases(
        capsys: CaptureFixture[str]) -> None:
    """Cell XML parsing keeps unsupported shared-string data unchanged."""
    assert number_from_cell_text('not-a-number') == 'not-a-number'
    assert number_from_cell_text('1e3') == 1000.0
    worksheet_xml = (
        '<?xml version="1.0" encoding="UTF-8"?>'
        '<worksheet '
        'xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main">'
        '<sheetData>'
        '<row r="1">'
        '<c><v>9</v></c>'
        '<c r="A1"/>'
        '<c r="A2" t="s"><v>bad-index</v></c>'
        '<c r="A3" t="s"><v>2</v></c>'
        '</row>'
        '</sheetData>'
        '</worksheet>'
    ).encode('utf-8')
    assert sheet_data_from_xml(worksheet_xml, {1: 'one'}, {}) == {
            'A2': {'v': 'bad-index', 'f': '', 's': '0', 'c': ''},
            'A3': {'v': '2', 'f': '', 's': '0', 'c': ''}
        }
    check_capsys(capsys)


def test_excel_pylightxl_named_ranges_and_typed_values_cover_style_paths(
        capsys: CaptureFixture[str]) -> None:
    """Named ranges and typed values handle the remaining style paths."""
    workbook_root = ET.fromstring(
        '<?xml version="1.0" encoding="UTF-8"?>'
        '<workbook '
        'xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main">'
        '<definedNames>'
        '<definedName>MissingName</definedName>'
        '<definedName name="MissingAddress"/>'
        '<definedName name="NoSheet">A1:B2</definedName>'
        '<definedName name="Totals">\'Sheet One\'!$A$1:$B$2</definedName>'
        '</definedNames>'
        '</workbook>')
    database = Database()
    database.add_ws('Sheet One', data={})
    load_named_ranges(workbook_root, database)
    assert database.nr_names == {'Totals': 'Sheet One!A1:B2'}
    assert style_index_for_code('14') == '1'
    assert style_index_for_code('18') == '2'
    assert style_index_for_code('22') == '3'
    assert style_index_for_code('999') is None
    when = datetime(2026, 3, 24, 14, 30, 0)
    excel_number = datetime_to_excel_number(when)
    assert _InspectableTableIOExcelPylightxl.parse_typed_value(
        '2026/03/24', '14') == datetime(2026, 3, 24)
    assert _InspectableTableIOExcelPylightxl.parse_typed_value(
        excel_number, '14') == datetime(2026, 3, 24)
    assert _InspectableTableIOExcelPylightxl.parse_typed_value(
        excel_number, '18') == '14:30:00'
    assert _InspectableTableIOExcelPylightxl.parse_typed_value(
        '2026/03/24 14:30:00', '22') == when
    check_capsys(capsys)


def test_excel_pylightxl_open_adds_default_sheet_for_empty_reader_result(
        monkeypatch: pytest.MonkeyPatch, capsys: CaptureFixture[str]) -> None:
    """READ open creates a default sheet when parsing yields no sheets."""
    with TemporaryDirectory() as temp_dir:
        Path(temp_dir, 'empty.xlsx').touch()

        def empty_database(_file_name: str) -> Database:
            """Return an empty database for the open fallback test."""
            return Database()

        monkeypatch.setattr(tableio_excel_pylightxl_module, '_read_database',
                            empty_database)
        table_io = _InspectableTableIOExcelPylightxl(Path(temp_dir) / 'empty',
                                                     FileAccess.READ)
        table_io.open()
        assert table_io.list_sheets() == ['Sheet1']
        assert table_io.run_last_used_column() == -1
        table_io.run_delete_filtered_range('ignored')
        table_io.run_add_filtered_range((0, 0, 1, 1), 'ignored')
        table_io.worksheet = object()
        assert table_io.run_current_sheet_name() == 'Sheet1'
        table_io.run_close_hook()
    check_capsys(capsys)


def test_excel_pylightxl_write_file_suffix_cleans_temp_file_on_error(
        monkeypatch: pytest.MonkeyPatch, capsys: CaptureFixture[str]) -> None:
    """Temporary workbook files are removed when rewrite cleanup fails."""
    with TemporaryDirectory() as temp_dir:
        table_io = _InspectableTableIOExcelPylightxl(
            Path(temp_dir) / 'write_error', FileAccess.CREATE)
        table_io.open()
        temp_path = Path(temp_dir) / 'temporary.xlsx'

        def raise_rewrite_error(_file_name: Path) -> None:
            """Raise the rewrite error used by this cleanup test."""
            raise RuntimeError('rewrite failed')

        monkeypatch.setattr(table_io, '_temporary_workbook_path',
                            lambda _source_path: temp_path)
        monkeypatch.setattr(table_io, '_rewrite_workbook_xml',
                            raise_rewrite_error)
        with pytest.raises(RuntimeError, match='rewrite failed'):
            table_io.run_write_file_suffix()
        assert not temp_path.exists()
        table_io.run_close_hook()
    check_capsys(capsys)


def test_excel_pylightxl_xml_rewrite_helpers_cover_existing_metadata(
        capsys: CaptureFixture[str]) -> None:
    """XML rewrite helpers handle existing styles and edge-case cells."""
    main_namespace = XML_NS['main']
    formula_cell = ET.fromstring(
        f'<c xmlns="{main_namespace}"><f>SUM(A1:A2)</f><v>None</v></c>')
    assert _InspectableTableIOExcelPylightxl.invalid_placeholder_cell(
        formula_cell) is False
    assert _InspectableTableIOExcelPylightxl.normalize_written_bool_cell(
        formula_cell) is False
    with TemporaryDirectory() as temp_dir:
        file_name = Path(temp_dir) / 'rewrite_helpers.xlsx'
        _write_zip_members(file_name, _rewrite_workbook_members())
        table_io = _InspectableTableIOExcelPylightxl(
            Path(temp_dir) / 'rewrite_helper_object', FileAccess.CREATE)
        table_io.set_sheet_style_codes({'sheet1': {'A3': '999'}})
        assert table_io.run_entry_style_codes(
            'xl/worksheets/missing.xml',
            {'Sheet1': (1, 'worksheets/sheet1.xml')}) == {}
        no_sheet_data = table_io.run_worksheet_xml_for_output(
            'xl/worksheets/missing.xml',
            (
                '<?xml version="1.0" encoding="UTF-8"?>'
                '<worksheet '
                'xmlns="http://schemas.openxmlformats.org/spreadsheetml/'
                '2006/main"/>'
            ).encode('utf-8'),
            {})
        assert b'mc:Ignorable=' in no_sheet_data
        table_io.run_rewrite_workbook_xml(file_name)
        with ZipFile(file_name) as zip_file:
            rewritten_worksheet_xml = zip_file.read(
                'xl/worksheets/sheet1.xml').decode('utf-8')
            rewritten_content_types = zip_file.read(
                '[Content_Types].xml').decode('utf-8')
            rewritten_workbook_rels = zip_file.read(
                'xl/_rels/workbook.xml.rels').decode('utf-8')
            rewritten_styles = zip_file.read('xl/styles.xml').decode('utf-8')
        assert 'old styles' not in rewritten_styles
        assert '<styleSheet' in rewritten_styles
        assert '<v>0</v>' in rewritten_worksheet_xml
        assert 'A1' not in rewritten_worksheet_xml
        assert 'spreadsheetml.styles+xml' in rewritten_content_types
        assert 'Target="styles.xml"' in rewritten_workbook_rels
    check_capsys(capsys)
