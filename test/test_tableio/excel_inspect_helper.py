#! /usr/local/bin/python3
"""Shared workbook inspection helpers for Excel backend tests."""

# Copyright (c) 2026 Tom Björkholm
# MIT License

from datetime import datetime
from pathlib import Path

import pytest
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet


def _cell_border_styles(
        worksheet: Worksheet, row_index: int,
        column_index: int) -> tuple[object, object, object, object]:
    """Return the saved border styles for one cell as top-right-bottom-left."""
    cell = worksheet.cell(row=row_index + 1, column=column_index + 1)
    return (cell.border.top.style, cell.border.right.style,
            cell.border.bottom.style, cell.border.left.style)


def inspect_formatted_workbook(file_path: Path) -> None:
    """Check formatting and filtered-range metadata in one workbook."""
    workbook = load_workbook(file_path)
    worksheet = workbook.active
    assert isinstance(worksheet, Worksheet)
    assert [table.ref for table in worksheet.tables.values()] == ['A1:B2']
    assert worksheet['A1'].font.bold is True
    assert worksheet['A2'].font.italic is True
    assert worksheet['A2'].fill.fill_type == 'solid'
    assert worksheet['A2'].fill.fgColor.rgb == 'FFFFFF00'
    workbook.close()


def inspect_multiple_filters_workbook(file_path: Path) -> None:
    """Check that sequential filtered writes remain separate tables."""
    workbook = load_workbook(file_path)
    worksheet = workbook.active
    assert isinstance(worksheet, Worksheet)
    assert sorted(table.ref for table in worksheet.tables.values()) == [
        'A1:B2',
        'A4:B5'
    ]
    workbook.close()


def inspect_table_width_cap_workbook(file_path: Path,
                                     expected_width: float) -> None:
    """Check that repeated boxed writes do not shrink the widened column."""
    workbook = load_workbook(file_path)
    worksheet = workbook.active
    assert isinstance(worksheet, Worksheet)
    assert worksheet['A2'].value == 'y'
    assert worksheet.column_dimensions['A'].width == pytest.approx(
        expected_width)
    workbook.close()


def inspect_rewrite_box_workbook(file_path: Path) -> None:
    """Check that overwriting a box removes stale table metadata."""
    workbook = load_workbook(file_path)
    worksheet = workbook.active
    assert isinstance(worksheet, Worksheet)
    assert not worksheet.tables
    workbook.close()


def inspect_row_formatted_workbook(file_path: Path) -> None:
    """Check row formatting copied from FmtDictRow writes."""
    workbook = load_workbook(file_path)
    worksheet = workbook.active
    assert isinstance(worksheet, Worksheet)
    assert [table.ref for table in worksheet.tables.values()] == ['A1:B3']
    assert worksheet['A2'].font.bold is True
    assert worksheet['B2'].font.bold is True
    assert worksheet['A2'].fill.fgColor.rgb == 'FFC6EFCE'
    assert worksheet['A3'].font.italic is True
    assert worksheet['B3'].font.italic is True
    assert worksheet['A3'].fill.fgColor.rgb == 'FFFFC7CE'
    workbook.close()


def inspect_dict_header_fmt_workbook(file_path: Path) -> None:
    """Check header formatting produced by dict-data writes."""
    workbook = load_workbook(file_path)
    worksheet = workbook.active
    assert isinstance(worksheet, Worksheet)
    assert worksheet['A1'].value == 'name'
    assert worksheet['B1'].value == 'active'
    assert worksheet['A1'].font.bold is True
    assert worksheet['B1'].font.bold is True
    assert worksheet['A1'].fill.fgColor.rgb == 'FFFFFF00'
    assert worksheet['B1'].fill.fgColor.rgb == 'FFFFFF00'
    assert worksheet['A2'].font.bold is False
    assert worksheet['B2'].font.bold is False
    workbook.close()


def inspect_fmtdict_header_fmt_workbook(file_path: Path) -> None:
    """Check header and row formatting separation in fmtdict writes."""
    workbook = load_workbook(file_path)
    worksheet = workbook.active
    assert isinstance(worksheet, Worksheet)
    assert worksheet['A1'].font.bold is True
    assert worksheet['B1'].font.bold is True
    assert worksheet['A2'].font.italic is True
    assert worksheet['B2'].font.italic is True
    assert worksheet['A2'].fill.fgColor.rgb == 'FFC6EFCE'
    assert worksheet['B2'].fill.fgColor.rgb == 'FFC6EFCE'
    workbook.close()


def inspect_table_width_heading_workbook(file_path: Path,
                                         expected_width_a: float,
                                         expected_width_b: float) -> None:
    """Check that heading text does not affect table column widths."""
    workbook = load_workbook(file_path)
    worksheet = workbook.active
    assert isinstance(worksheet, Worksheet)
    assert worksheet.column_dimensions['A'].width == pytest.approx(
        expected_width_a)
    assert worksheet.column_dimensions['B'].width == pytest.approx(
        expected_width_b)
    workbook.close()


def inspect_normalized_header_workbook(
        file_path: Path, sheet_name: str = 'Sheet1') -> None:
    """Check one workbook with normalized filtered table headers."""
    workbook = load_workbook(file_path)
    worksheet = workbook[sheet_name]
    assert isinstance(worksheet, Worksheet)
    assert worksheet['A1'].value == '1'
    assert worksheet['B1'].value == 'Column2'
    assert [table.ref for table in worksheet.tables.values()] == [
        'A1:B2'
    ]
    workbook.close()


def inspect_datetime_cells_workbook(file_path: Path) -> None:
    """Check that one written workbook stores datetime cells as datetimes."""
    workbook = load_workbook(file_path, data_only=True)
    worksheet = workbook.active
    assert isinstance(worksheet, Worksheet)
    assert worksheet['A2'].value is True
    assert worksheet['B2'].value == datetime(2026, 3, 24, 14, 30, 0)
    workbook.close()


def inspect_bordered_workbook(file_path: Path) -> None:
    """Check that one workbook stores the requested table borders."""
    workbook = load_workbook(file_path)
    worksheet = workbook.active
    assert isinstance(worksheet, Worksheet)
    assert _cell_border_styles(worksheet, 0, 0) == (
        'medium', 'thin', 'thin', 'medium')
    assert _cell_border_styles(worksheet, 0, 1) == (
        'medium', 'medium', 'thin', 'thin')
    assert _cell_border_styles(worksheet, 1, 0) == (
        'thin', 'thin', 'medium', 'medium')
    assert _cell_border_styles(worksheet, 1, 1) == (
        'thin', 'medium', 'medium', 'thin')
    workbook.close()


def inspect_box_rewrite_clears_borders_workbook(file_path: Path) -> None:
    """Check that boxed rewrites clear stale cell borders."""
    workbook = load_workbook(file_path)
    worksheet = workbook.active
    assert isinstance(worksheet, Worksheet)
    for row_index in range(2):
        for column_index in range(2):
            assert _cell_border_styles(worksheet, row_index, column_index) == (
                None, None, None, None)
    workbook.close()
