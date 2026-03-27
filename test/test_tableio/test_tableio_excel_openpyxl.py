#! /usr/local/bin/python3
"""Tests for the tableio_excel_openpyxl module."""

# Copyright (c) 2026 Tom Björkholm
# MIT License

from datetime import datetime
from pathlib import Path
from tempfile import TemporaryDirectory
from typing import Optional
from xml.etree import ElementTree as ET
from zipfile import ZipFile

from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from pytest import CaptureFixture

from tableio.tableio import FileAccess
from tableio.tableio_excel_openpyxl import TableIOExcelOpenPyXL

from .check_capsys import check_capsys
from .spreadsheet_test_helper import \
    run_box_write_removes_overlapping_filtered_range, \
    run_multi_sheet_heading_state_is_per_sheet, \
    run_multi_sheet_read_only_create_raises, \
    run_multi_sheet_read_positions_are_per_sheet, \
    run_multi_sheet_update_uses_selected_sheet_write_position, \
    run_multi_sheet_write_positions_are_per_sheet, \
    run_read_formula_uses_cached_value, \
    run_read_formula_without_cached_value, \
    run_round_trip_dictdata_in_box, \
    run_round_trip_sequential_list_reads, \
    run_table_width_is_widen_only_with_cap, \
    run_update_default_write_starts_after_last_used_row, \
    run_write_dictdata_applies_first_row_format, \
    run_write_fmtdictdata_applies_first_row_format, \
    run_write_formatted_listdata_applies_formatting_and_filter, \
    run_write_multiple_filtered_ranges_keeps_all_ranges, \
    run_write_row_formatted_dictdata_applies_formatting


_XML_NS = {
    'sheet': 'http://schemas.openxmlformats.org/spreadsheetml/2006/main'
}


def _create_formula_workbook(file_path: Path,
                             cached_value: Optional[int] = None) -> None:
    """Create a workbook with one formula cell and an optional cached value."""
    workbook = Workbook()
    worksheet = workbook.active
    assert isinstance(worksheet, Worksheet)
    worksheet['A1'] = '=1+2'
    worksheet['B1'] = 'x'
    workbook.save(file_path)
    workbook.close()
    if cached_value is None:
        return
    temp_path = file_path.with_name(f'{file_path.stem}_tmp.xlsx')
    with ZipFile(file_path) as zip_file, \
            ZipFile(temp_path, 'w') as temp_zip:
        for item in zip_file.infolist():
            data = zip_file.read(item.filename)
            if item.filename == 'xl/worksheets/sheet1.xml':
                root = ET.fromstring(data)
                cell = root.find('.//sheet:c[@r="A1"]', _XML_NS)
                assert cell is not None
                value = cell.find('sheet:v', _XML_NS)
                assert value is not None
                value.text = str(cached_value)
                data = ET.tostring(root, encoding='utf-8',
                                   xml_declaration=False)
            temp_zip.writestr(item, data)
    temp_path.replace(file_path)


def _create_update_workbook(file_path: Path) -> None:
    """Create the starting workbook used by the UPDATE mode test."""
    workbook = Workbook()
    worksheet = workbook.active
    assert isinstance(worksheet, Worksheet)
    worksheet['A1'] = 'old'
    worksheet['B1'] = 'row'
    workbook.save(file_path)
    workbook.close()


def _inspect_updated_workbook(file_path: Path) -> None:
    """Check the workbook produced by the shared UPDATE mode case."""
    workbook = load_workbook(file_path, data_only=True)
    worksheet = workbook.active
    assert isinstance(worksheet, Worksheet)
    assert worksheet['A1'].value == 'old'
    assert worksheet['B1'].value == 'row'
    assert worksheet['A2'].value is None
    assert worksheet['B2'].value is None
    assert worksheet['A3'].value == 'new'
    assert worksheet['B3'].value == 'row'
    workbook.close()


def _inspect_formatted_workbook(file_path: Path) -> None:
    """Check formatting and filtered-range metadata in one workbook."""
    workbook = load_workbook(file_path)
    worksheet = workbook.active
    assert isinstance(worksheet, Worksheet)
    assert [table.ref for table in worksheet.tables.values()] == ['A1:B2']
    assert worksheet['A1'].font.bold is True
    assert worksheet['A2'].font.italic is True
    assert worksheet['A2'].fill.fill_type == 'solid'
    assert worksheet['A2'].fill.fgColor.rgb == 'FFFFFF00'
    workbook.close()


def _inspect_multiple_filters_workbook(file_path: Path) -> None:
    """Check that sequential filtered writes remain separate tables."""
    workbook = load_workbook(file_path)
    worksheet = workbook.active
    assert isinstance(worksheet, Worksheet)
    assert sorted(table.ref for table in worksheet.tables.values()) == [
        'A1:B2',
        'A4:B5'
    ]
    workbook.close()


def _inspect_table_width_cap_workbook(file_path: Path) -> None:
    """Check that repeated boxed writes do not shrink the widened column."""
    workbook = load_workbook(file_path)
    worksheet = workbook.active
    assert isinstance(worksheet, Worksheet)
    assert worksheet['A2'].value == 'y'
    assert worksheet.column_dimensions['A'].width == 50.0
    workbook.close()


def _inspect_rewrite_box_workbook(file_path: Path) -> None:
    """Check that overwriting a box removes stale table metadata."""
    workbook = load_workbook(file_path)
    worksheet = workbook.active
    assert isinstance(worksheet, Worksheet)
    assert not worksheet.tables
    workbook.close()


def _inspect_row_formatted_workbook(file_path: Path) -> None:
    """Check row formatting copied from FmtDictRow writes."""
    workbook = load_workbook(file_path)
    worksheet = workbook.active
    assert isinstance(worksheet, Worksheet)
    assert [table.ref for table in worksheet.tables.values()] == ['A1:B3']
    assert worksheet['A2'].font.bold is True
    assert worksheet['B2'].font.bold is True
    assert worksheet['A2'].fill.fgColor.rgb == 'FFC6EFCE'
    assert worksheet['A3'].font.italic is True
    assert worksheet['B3'].font.italic is True
    assert worksheet['A3'].fill.fgColor.rgb == 'FFFFC7CE'
    workbook.close()


def _inspect_dict_header_fmt_workbook(file_path: Path) -> None:
    """Check header formatting produced by dict-data writes."""
    workbook = load_workbook(file_path)
    worksheet = workbook.active
    assert isinstance(worksheet, Worksheet)
    assert worksheet['A1'].value == 'name'
    assert worksheet['B1'].value == 'active'
    assert worksheet['A1'].font.bold is True
    assert worksheet['B1'].font.bold is True
    assert worksheet['A1'].fill.fgColor.rgb == 'FFFFFF00'
    assert worksheet['B1'].fill.fgColor.rgb == 'FFFFFF00'
    assert worksheet['A2'].font.bold is False
    assert worksheet['B2'].font.bold is False
    workbook.close()


def _inspect_fmtdict_header_fmt_workbook(file_path: Path) -> None:
    """Check header and row formatting separation in fmtdict writes."""
    workbook = load_workbook(file_path)
    worksheet = workbook.active
    assert isinstance(worksheet, Worksheet)
    assert worksheet['A1'].font.bold is True
    assert worksheet['B1'].font.bold is True
    assert worksheet['A2'].font.italic is True
    assert worksheet['B2'].font.italic is True
    assert worksheet['A2'].fill.fgColor.rgb == 'FFC6EFCE'
    assert worksheet['B2'].fill.fgColor.rgb == 'FFC6EFCE'
    workbook.close()


def test_excel_round_trip_sequential_list_reads(
        capsys: CaptureFixture[str]) -> None:
    """Two list sections can be written and then read back sequentially."""
    run_round_trip_sequential_list_reads(TableIOExcelOpenPyXL, capsys)


def test_excel_round_trip_dictdata_in_box(
        capsys: CaptureFixture[str]) -> None:
    """Dict data can be written into and read back from a box."""
    run_round_trip_dictdata_in_box(TableIOExcelOpenPyXL, capsys)


def test_excel_multi_sheet_write_positions_are_per_sheet(
        capsys: CaptureFixture[str]) -> None:
    """Sequential writes keep an independent default position per sheet."""
    run_multi_sheet_write_positions_are_per_sheet(
        TableIOExcelOpenPyXL, capsys)


def test_excel_multi_sheet_read_positions_are_per_sheet(
        capsys: CaptureFixture[str]) -> None:
    """Sequential reads resume independently when switching sheets."""
    run_multi_sheet_read_positions_are_per_sheet(
        TableIOExcelOpenPyXL, capsys)


def test_excel_multi_sheet_heading_state_is_per_sheet(
        capsys: CaptureFixture[str]) -> None:
    """Each sheet tracks whether a default heading level was used before."""
    run_multi_sheet_heading_state_is_per_sheet(
        TableIOExcelOpenPyXL, capsys)


def test_excel_multi_sheet_read_only_create_raises(
        capsys: CaptureFixture[str]) -> None:
    """READ mode can select an existing sheet but cannot create one."""
    run_multi_sheet_read_only_create_raises(
        TableIOExcelOpenPyXL, capsys)


def test_excel_update_default_write_starts_after_last_used_row(
        capsys: CaptureFixture[str]) -> None:
    """UPDATE mode appends after the used area with a blank row separator."""
    run_update_default_write_starts_after_last_used_row(
        TableIOExcelOpenPyXL, '.xlsx', _create_update_workbook,
        _inspect_updated_workbook, capsys)


def test_excel_multi_sheet_update_uses_selected_sheet_write_position(
        capsys: CaptureFixture[str]) -> None:
    """UPDATE mode appends using the selected sheet's used area."""
    run_multi_sheet_update_uses_selected_sheet_write_position(
        TableIOExcelOpenPyXL, capsys)


def test_excel_write_formatted_listdata_applies_formatting_and_filter(
        capsys: CaptureFixture[str]) -> None:
    """Per-cell formatting and one filtered table are written."""
    run_write_formatted_listdata_applies_formatting_and_filter(
        TableIOExcelOpenPyXL, '.xlsx', _inspect_formatted_workbook,
        capsys)


def test_excel_table_width_uses_table_content_not_heading(
        capsys: CaptureFixture[str]) -> None:
    """Table column widths ignore headings written outside table cells."""
    with TemporaryDirectory() as temp_dir:
        file_name = Path(temp_dir) / 'table_width_heading'
        with TableIOExcelOpenPyXL(file_name, FileAccess.CREATE) as table_io:
            table_io.write_heading('Heading text that is much wider than '
                                   'column A needs')
            table_io.write_table_listdata([
                ['id', 'report date'],
                ['A', datetime(2026, 3, 24, 14, 30, 0)]
            ])
        workbook = load_workbook(Path(temp_dir) / 'table_width_heading.xlsx')
        worksheet = workbook.active
        assert isinstance(worksheet, Worksheet)
        assert worksheet.column_dimensions['A'].width == 13.0
        assert worksheet.column_dimensions['B'].width == 21.0
        workbook.close()
    check_capsys(capsys)


def test_excel_write_multiple_filtered_ranges_keeps_all_tables(
        capsys: CaptureFixture[str]) -> None:
    """Sequential filtered writes are kept as separate worksheet tables."""
    run_write_multiple_filtered_ranges_keeps_all_ranges(
        TableIOExcelOpenPyXL, '.xlsx',
        _inspect_multiple_filters_workbook, capsys)


def test_excel_table_width_is_widen_only_with_cap(
        capsys: CaptureFixture[str]) -> None:
    """Box rewrites keep an already widened column width."""
    run_table_width_is_widen_only_with_cap(
        TableIOExcelOpenPyXL, '.xlsx',
        _inspect_table_width_cap_workbook, capsys)


def test_excel_box_write_removes_overlapping_filtered_table(
        capsys: CaptureFixture[str]) -> None:
    """Rewriting a boxed area removes any stale overlapping table metadata."""
    run_box_write_removes_overlapping_filtered_range(
        TableIOExcelOpenPyXL, '.xlsx',
        _inspect_rewrite_box_workbook, capsys)


def test_excel_write_row_formatted_dictdata_applies_formatting(
        capsys: CaptureFixture[str]) -> None:
    """Row formatting for dict rows is copied to each written cell."""
    run_write_row_formatted_dictdata_applies_formatting(
        TableIOExcelOpenPyXL, '.xlsx',
        _inspect_row_formatted_workbook, capsys)


def test_excel_write_dictdata_applies_first_row_format(
        capsys: CaptureFixture[str]) -> None:
    """Dict header cells can be formatted with first_row_format."""
    run_write_dictdata_applies_first_row_format(
        TableIOExcelOpenPyXL, '.xlsx',
        _inspect_dict_header_fmt_workbook, capsys)


def test_excel_write_fmtdictdata_applies_first_row_format(
        capsys: CaptureFixture[str]) -> None:
    """Formatted dict writes keep header and data-row formatting separate."""
    run_write_fmtdictdata_applies_first_row_format(
        TableIOExcelOpenPyXL, '.xlsx',
        _inspect_fmtdict_header_fmt_workbook, capsys)


def test_excel_read_formula_uses_cached_value(
        capsys: CaptureFixture[str]) -> None:
    """A formula cell is read as its cached value."""
    run_read_formula_uses_cached_value(
        TableIOExcelOpenPyXL, '.xlsx', _create_formula_workbook, 3,
        capsys)


def test_excel_read_formula_without_cached_value_returns_none(
        capsys: CaptureFixture[str]) -> None:
    """A formula without a cached result is read as None."""
    run_read_formula_without_cached_value(
        TableIOExcelOpenPyXL, '.xlsx', _create_formula_workbook,
        capsys)
