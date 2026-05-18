#! /usr/local/bin/python3
"""Tests for the tableio_excel_xlsxwriter module."""

# Copyright (c) 2026 Tom Björkholm
# MIT License

import io
from datetime import datetime
from pathlib import Path
from tempfile import TemporaryDirectory
from typing import Optional

import pytest
from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from openxml_audit import OpenXmlValidator  # type: ignore[import-untyped]
from pytest import CaptureFixture

from tableio.capability import CAP_IMPLEMENTED, CAP_UNSUPPORTED, \
    Capabilities, CapabilityNotSupported
from tableio.border_helper import CellStyleState, DEFAULT_CELL_STYLE
from tableio.factory import create_tableio
from tableio.color import Color
from tableio.tableio import Box, FileAccess
from tableio.tableio_excel_openpyxl import TableIOExcelOpenPyXL
from tableio.tableio_excel_xlsxwriter import TableIOExcelXlsxWriter
from tableio.value_type import Fmt, ValueFmt

from .check_capsys import check_capsys
from .excel_inspect_helper import inspect_bordered_workbook, \
    inspect_box_rewrite_clears_borders_workbook, \
    inspect_datetime_cells_workbook, inspect_dict_header_fmt_workbook, \
    inspect_fmtdict_header_fmt_workbook, \
    inspect_formatted_workbook, inspect_multiple_filters_workbook, \
    inspect_normalized_header_workbook, inspect_rewrite_box_workbook, \
    inspect_row_formatted_workbook, inspect_table_width_cap_workbook, \
    inspect_table_width_heading_workbook
from .spreadsheet_test_helper import \
    run_bordered_workbook_is_validator_clean, \
    run_box_rewrite_clears_borders, \
    run_box_write_removes_overlapping_filtered_range, \
    run_boxed_table_partial_overwrite_raises, \
    run_multi_sheet_heading_state_is_per_sheet, \
    run_multi_sheet_write_positions_are_per_sheet, \
    run_open_rejects_second_open, \
    run_select_missing_sheet_without_create_raises_key_error, \
    run_table_width_uses_table_content_not_heading, \
    run_table_width_is_widen_only_with_cap, \
    run_write_dictdata_applies_first_row_format, \
    run_write_fmtdictdata_applies_first_row_format, \
    run_write_formatted_listdata_applies_formatting_and_filter, \
    run_write_table_listdata_applies_borders, \
    run_write_multiple_filtered_ranges_keeps_all_ranges, \
    run_write_row_formatted_dictdata_applies_formatting


_XLSXWRITER_WIDTH_OFFSET = 0.7109375


class _FailingWorksheet:
    """Worksheet double whose add_table call always reports failure."""

    def __init__(self) -> None:
        """Initialize the minimal worksheet state used by the tests."""
        self.tables: list[dict[str, object]] = []
        self.table_cells: dict[tuple[int, int], str] = {}
        self.filter_cells: dict[tuple[int, int], tuple[str, str]] = {}

    def add_table(self, *args: object, **kwargs: object) -> int:
        """Simulate a backend failure when adding a table."""
        _ = args
        _ = kwargs
        return 1

    def set_column(self, *args: object, **kwargs: object) -> object:
        """Ignore width requests because the test never inspects them."""
        _ = args
        _ = kwargs
        return 0

    def write(self, row: int, col: int, *args: object) -> object:
        """Ignore cell writes because the failure happens earlier."""
        _ = row
        _ = col
        _ = args
        return 0

    def write_blank(self, row: int, col: int, blank: object,
                    cell_format: Optional[object] = None) -> object:
        """Ignore blank writes because the failure happens earlier."""
        _ = row
        _ = col
        _ = blank
        _ = cell_format
        return 0


class _InspectableTableIOExcelXlsxWriter(TableIOExcelXlsxWriter):
    """Expose protected XlsxWriter helpers needed by the tests."""

    def run_write_file_suffix(self) -> None:
        """Expose the workbook-finalize hook for tests."""
        self._write_file_suffix()

    def run_last_used_column(self) -> int:
        """Expose the empty-sheet column helper for tests."""
        return self._last_used_column(self._write_sheet())

    def run_add_filtered_range(self, bounds: tuple[int, int, int, int],
                               name: str) -> None:
        """Expose filtered-range creation for tests."""
        self._add_filtered_range(bounds, name)

    def run_xlsx_format(self, style: Optional[CellStyleState],
                        datetime_value: bool) -> Optional[object]:
        """Expose the XlsxWriter format-cache helper for tests."""
        return self._xlsx_format(style, datetime_value)


def test_excel_xlsxwriter_get_capabilities(
        capsys: CaptureFixture[str]) -> None:
    """The backend reports the honest XlsxWriter feature set."""
    capabilities = TableIOExcelXlsxWriter.get_capabilities()
    assert capabilities.can_read == CAP_UNSUPPORTED
    assert capabilities.can_write == CAP_IMPLEMENTED
    assert capabilities.can_fmt_row == CAP_IMPLEMENTED
    assert capabilities.can_fmt_value == CAP_IMPLEMENTED
    assert capabilities.filtered_data_range == CAP_IMPLEMENTED
    assert capabilities.can_write_box == CAP_IMPLEMENTED
    assert capabilities.can_read_box == CAP_UNSUPPORTED
    assert capabilities.can_write_highlight == CAP_IMPLEMENTED
    assert capabilities.multi_sheet == CAP_IMPLEMENTED
    assert capabilities.can_find_value_position == CAP_UNSUPPORTED
    assert capabilities.can_write_borders == CAP_IMPLEMENTED
    check_capsys(capsys)


def test_excel_xlsxwriter_get_description_prefers_write_backend(
        capsys: CaptureFixture[str]) -> None:
    """The XlsxWriter descriptor prefers it for write-only Excel output."""
    description = TableIOExcelXlsxWriter.get_description()
    assert description.format_name == 'Excel'
    assert description.implementation == 'XlsxWriter'
    openpyxl_description = TableIOExcelOpenPyXL.get_description()
    assert description.priority > openpyxl_description.priority
    check_capsys(capsys)


def test_excel_factory_prefers_xlsxwriter_for_create(
        capsys: CaptureFixture[str]) -> None:
    """Factory selection uses XlsxWriter by default for CREATE access."""
    with TemporaryDirectory() as temp_dir:
        table_io = create_tableio('Excel', Path(temp_dir) / 'create_target',
                                  FileAccess.CREATE)
        assert isinstance(table_io, TableIOExcelXlsxWriter)
    check_capsys(capsys)


def test_excel_factory_prefers_openpyxl_for_read_and_update(
        capsys: CaptureFixture[str]) -> None:
    """Explicit read/write capabilities select the readable backend."""
    with TemporaryDirectory() as temp_dir:
        workbook = Workbook()
        workbook.save(Path(temp_dir) / 'existing.xlsx')
        workbook.close()
        read_table_io = create_tableio('Excel', Path(temp_dir) / 'existing',
                                       FileAccess.READ,
                                       capabilities=Capabilities(
                                           can_read=CAP_IMPLEMENTED))
        update_table_io = create_tableio('Excel', Path(temp_dir) / 'existing',
                                         FileAccess.UPDATE,
                                         capabilities=Capabilities(
                                             can_read=CAP_IMPLEMENTED,
                                             can_write=CAP_IMPLEMENTED))
        assert isinstance(read_table_io, TableIOExcelOpenPyXL)
        assert isinstance(update_table_io, TableIOExcelOpenPyXL)
    check_capsys(capsys)


@pytest.mark.parametrize('file_access', [FileAccess.READ, FileAccess.UPDATE])
def test_excel_xlsxwriter_open_rejects_read_and_update(
        file_access: FileAccess, capsys: CaptureFixture[str]) -> None:
    """Reject access modes that need existing workbook reads."""
    with TemporaryDirectory() as temp_dir:
        workbook = Workbook()
        workbook.save(Path(temp_dir) / 'existing.xlsx')
        workbook.close()
        table_io = TableIOExcelXlsxWriter(Path(temp_dir) / 'existing',
                                          file_access)
        with pytest.raises(io.UnsupportedOperation,
                           match='only create new Excel files'):
            table_io.open()
    check_capsys(capsys)


def test_excel_xlsxwriter_read_methods_raise(
        capsys: CaptureFixture[str]) -> None:
    """Read APIs stay unavailable even during a CREATE session."""
    with TemporaryDirectory() as temp_dir:
        with TableIOExcelXlsxWriter(Path(temp_dir) / 'write_only',
                                    FileAccess.CREATE) as table_io:
            table_io.write_table_listdata([['name', 'active'],
                                           ['Alice', True]])
            with pytest.raises(CapabilityNotSupported,
                               match='read table data'):
                table_io.read_table_listdata()
            with pytest.raises(CapabilityNotSupported,
                               match='read table data'):
                table_io.read_table_dictdata()
    check_capsys(capsys)


def test_excel_xlsxwriter_find_value_raises(
        capsys: CaptureFixture[str]) -> None:
    """The backend rejects find_value since it cannot read worksheets."""
    with TemporaryDirectory() as temp_dir:
        with TableIOExcelXlsxWriter(Path(temp_dir) / 'find_value',
                                    FileAccess.CREATE) as table_io:
            table_io.write_table_listdata([['name', 'active'],
                                           ['Alice', True]])
            with pytest.raises(CapabilityNotSupported,
                               match='Finding value position is not '
                               'supported'):
                table_io.find_value('Alice')
    check_capsys(capsys)


def test_excel_xlsxwriter_read_cells_raises(
        capsys: CaptureFixture[str]) -> None:
    """The backend rejects reading a boxed cell range."""
    with TemporaryDirectory() as temp_dir:
        with TableIOExcelXlsxWriter(Path(temp_dir) / 'read_cells',
                                    FileAccess.CREATE) as table_io:
            table_io.write_table_listdata([['name', 'active'],
                                           ['Alice', True]])
            with pytest.raises(CapabilityNotSupported,
                               match='Reading from a box is not supported'):
                table_io.read_cells(Box(top=0, left=0, bottom=1, right=2))
    check_capsys(capsys)


def test_excel_xlsxwriter_write_formatted_listdata_applies_filter(
        capsys: CaptureFixture[str]) -> None:
    """Per-cell formatting and one filtered table are written."""
    run_write_formatted_listdata_applies_formatting_and_filter(
        TableIOExcelXlsxWriter, '.xlsx', inspect_formatted_workbook, capsys)


def test_excel_xlsxwriter_table_width_uses_table_content_not_heading(
        capsys: CaptureFixture[str]) -> None:
    """Table column widths ignore headings written outside table cells."""
    run_table_width_uses_table_content_not_heading(
        TableIOExcelXlsxWriter, '.xlsx',
        lambda file_path: inspect_table_width_heading_workbook(
            file_path, 13.0, 21.0 + _XLSXWRITER_WIDTH_OFFSET), capsys)


def test_excel_xlsxwriter_write_multiple_filtered_ranges_keeps_all_tables(
        capsys: CaptureFixture[str]) -> None:
    """Sequential filtered writes are kept as separate worksheet tables."""
    inspector = inspect_multiple_filters_workbook
    run_write_multiple_filtered_ranges_keeps_all_ranges(TableIOExcelXlsxWriter,
                                                        '.xlsx', inspector,
                                                        capsys)


def test_excel_xlsxwriter_table_width_is_widen_only_with_cap(
        capsys: CaptureFixture[str]) -> None:
    """Box rewrites keep an already widened column width."""
    run_table_width_is_widen_only_with_cap(
        TableIOExcelXlsxWriter, '.xlsx',
        lambda file_path: inspect_table_width_cap_workbook(
            file_path, 50.0 + _XLSXWRITER_WIDTH_OFFSET), capsys)


def test_excel_xlsxwriter_box_write_removes_overlapping_filtered_table(
        capsys: CaptureFixture[str]) -> None:
    """Rewriting a boxed area removes any stale overlapping table metadata."""
    run_box_write_removes_overlapping_filtered_range(
        TableIOExcelXlsxWriter, '.xlsx', inspect_rewrite_box_workbook, capsys)


def test_excel_xlsxwriter_boxed_table_partial_overwrite_raises(
        capsys: CaptureFixture[str]) -> None:
    """Boxed table writes reject overlaps that leave part of a table behind."""
    run_boxed_table_partial_overwrite_raises(TableIOExcelXlsxWriter, capsys)


def test_excel_xlsxwriter_multi_sheet_write_positions_are_per_sheet(
        capsys: CaptureFixture[str]) -> None:
    """Sequential writes keep an independent default position per sheet."""
    run_multi_sheet_write_positions_are_per_sheet(
        TableIOExcelXlsxWriter, capsys)


def test_excel_xlsxwriter_multi_sheet_heading_state_is_per_sheet(
        capsys: CaptureFixture[str]) -> None:
    """Each sheet tracks whether a default heading level was used before."""
    run_multi_sheet_heading_state_is_per_sheet(TableIOExcelXlsxWriter, capsys)


def test_excel_xlsxwriter_open_rejects_second_open(
        capsys: CaptureFixture[str]) -> None:
    """Opening the same Excel object twice raises RuntimeError."""
    run_open_rejects_second_open(TableIOExcelXlsxWriter, capsys)


def test_excel_xlsxwriter_select_missing_sheet_without_create_raises(
        capsys: CaptureFixture[str]) -> None:
    """Selecting a missing sheet without create=True raises KeyError."""
    run_select_missing_sheet_without_create_raises_key_error(
        TableIOExcelXlsxWriter, capsys)


def test_excel_xlsxwriter_write_row_formatted_dictdata_applies_formatting(
        capsys: CaptureFixture[str]) -> None:
    """Row formatting for dict rows is copied to each written cell."""
    inspector = inspect_row_formatted_workbook
    run_write_row_formatted_dictdata_applies_formatting(TableIOExcelXlsxWriter,
                                                        '.xlsx', inspector,
                                                        capsys)


def test_excel_xlsxwriter_write_dictdata_applies_first_row_format(
        capsys: CaptureFixture[str]) -> None:
    """Dict header cells can be formatted with first_row_format."""
    inspector = inspect_dict_header_fmt_workbook
    run_write_dictdata_applies_first_row_format(TableIOExcelXlsxWriter,
                                                '.xlsx', inspector, capsys)


def test_excel_xlsxwriter_write_fmtdictdata_applies_first_row_format(
        capsys: CaptureFixture[str]) -> None:
    """Formatted dict writes keep header and data-row formatting separate."""
    inspector = inspect_fmtdict_header_fmt_workbook
    run_write_fmtdictdata_applies_first_row_format(TableIOExcelXlsxWriter,
                                                   '.xlsx', inspector, capsys)


def test_excel_xlsxwriter_write_table_listdata_applies_borders(
        capsys: CaptureFixture[str]) -> None:
    """Writes the requested table borders to saved XlsxWriter cells."""
    run_write_table_listdata_applies_borders(TableIOExcelXlsxWriter, '.xlsx',
                                             inspect_bordered_workbook, capsys)


def test_excel_xlsxwriter_box_rewrite_clears_old_borders(
        capsys: CaptureFixture[str]) -> None:
    """Rewriting the same boxed area clears any stale cell borders."""
    run_box_rewrite_clears_borders(
        TableIOExcelXlsxWriter, '.xlsx',
        inspect_box_rewrite_clears_borders_workbook, capsys)


def test_excel_xlsxwriter_filtered_table_headers_are_normalized(
        capsys: CaptureFixture[str]) -> None:
    """Filtered table headers are normalized to valid Excel strings."""
    with TemporaryDirectory() as temp_dir:
        file_name = Path(temp_dir) / 'normalized_headers'
        with TableIOExcelXlsxWriter(file_name, FileAccess.CREATE) as table_io:
            table_io.write_table_listdata([[1, None], [2, 3]],
                                          filtered_data_range=True)
        inspect_normalized_header_workbook(
            Path(temp_dir) / 'normalized_headers.xlsx')
    check_capsys(capsys)


def test_excel_xlsxwriter_datetime_cells_round_trip_as_datetimes(
        capsys: CaptureFixture[str]) -> None:
    """Written datetime cells are stored with an Excel datetime format."""
    with TemporaryDirectory() as temp_dir:
        file_name = Path(temp_dir) / 'datetime_cells'
        with TableIOExcelXlsxWriter(file_name, FileAccess.CREATE) as table_io:
            table_io.write_table_listdata([
                ['flag', 'when'],
                [True, datetime(2026, 3, 24, 14, 30, 0)]
            ])
        inspect_datetime_cells_workbook(Path(temp_dir) / 'datetime_cells.xlsx')
    check_capsys(capsys)


def test_excel_xlsxwriter_write_cells_writes_exact_box_and_formatting(
        capsys: CaptureFixture[str]) -> None:
    """The write_cells method writes to the requested box."""
    with TemporaryDirectory() as temp_dir:
        file_name = Path(temp_dir) / 'write_cells'
        table_io = TableIOExcelXlsxWriter(file_name, FileAccess.CREATE)
        with table_io:
            table_io.write_table_listdata([['name', 'active'],
                                           ['Alice', True]])
            read_row = table_io.read_row
            write_row = table_io.write_row
            table_io.write_cells([
                [ValueFmt(value='Bob', fmt=Fmt(highlight=Color.YELLOW)),
                 ValueFmt(value=False, fmt=Fmt(bold=True))]
            ], Box(top=3, left=1, bottom=4, right=3))
            assert table_io.read_row == read_row
            assert table_io.write_row == write_row
        workbook = load_workbook(Path(temp_dir) / 'write_cells.xlsx')
        worksheet = workbook.active
        assert isinstance(worksheet, Worksheet)
        assert worksheet['B4'].value == 'Bob'
        assert worksheet['C4'].value is False
        assert worksheet['B4'].fill.fgColor.rgb == 'FFFFFF00'
        assert worksheet['C4'].font.bold is True
        workbook.close()
    check_capsys(capsys)


def test_excel_xlsxwriter_write_file_suffix_is_noop_without_workbook(
        capsys: CaptureFixture[str]) -> None:
    """Closing an unopened writer leaves the target file untouched."""
    with TemporaryDirectory() as temp_dir:
        table_io = _InspectableTableIOExcelXlsxWriter(
            Path(temp_dir) / 'no_workbook', FileAccess.CREATE)
        table_io.run_write_file_suffix()
        assert not Path(temp_dir, 'no_workbook.xlsx').exists()
    check_capsys(capsys)


def test_excel_xlsxwriter_default_cell_style_needs_no_format(
        capsys: CaptureFixture[str]) -> None:
    """The default style object is treated the same as no style."""
    table_io = _InspectableTableIOExcelXlsxWriter(Path('default_style'),
                                                  FileAccess.CREATE)
    assert table_io.run_xlsx_format(DEFAULT_CELL_STYLE, False) is None
    check_capsys(capsys)


def test_excel_xlsxwriter_last_used_column_is_minus_one_on_empty_sheet(
        capsys: CaptureFixture[str]) -> None:
    """Empty worksheets report no used column."""
    with TemporaryDirectory() as temp_dir:
        table_io = _InspectableTableIOExcelXlsxWriter(
            Path(temp_dir) / 'empty_sheet', FileAccess.CREATE)
        with table_io:
            assert table_io.run_last_used_column() == -1
    check_capsys(capsys)


def test_excel_xlsxwriter_filtered_range_raises_on_table_failure(
        capsys: CaptureFixture[str]) -> None:
    """Filtered range creation raises when XlsxWriter rejects the table."""
    with TemporaryDirectory() as temp_dir:
        table_io = _InspectableTableIOExcelXlsxWriter(
            Path(temp_dir) / 'table_failure', FileAccess.CREATE)
        with table_io:
            assert table_io.sheet_state is not None
            table_io.sheet_state.values[(0, 0)] = 'name'
            table_io.sheet_state.values[(0, 1)] = 'active'
            table_io.sheet_state.worksheet = _FailingWorksheet()
            with pytest.raises(ValueError,
                               match='Unable to create filtered Excel table'):
                table_io.run_add_filtered_range((0, 0, 2, 2), 'BrokenFilter')
    check_capsys(capsys)


def test_excel_xlsxwriter_bordered_workbook_is_validator_clean() -> None:
    """Bordered tables are written to validator-clean `.xlsx` files."""
    run_bordered_workbook_is_validator_clean(
        TableIOExcelXlsxWriter, '.xlsx',
        lambda file_path: OpenXmlValidator().validate(file_path).is_valid)
