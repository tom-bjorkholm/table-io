#! /usr/local/bin/python3
"""Shared workbook fixture helpers for Excel backend tests."""

# Copyright (c) 2026 Tom Björkholm
# MIT License

from pathlib import Path
from typing import Optional
from xml.etree import ElementTree as ET
from zipfile import ZipFile

from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet


_XML_NS = {
    'sheet': 'http://schemas.openxmlformats.org/spreadsheetml/2006/main'
}


def create_formula_workbook(file_path: Path,
                            cached_value: Optional[int] = None) -> None:
    """Create a workbook with one formula cell and an optional cached value."""
    workbook = Workbook()
    worksheet = workbook.active
    assert isinstance(worksheet, Worksheet)
    worksheet['A1'] = '=1+2'
    worksheet['B1'] = 'x'
    workbook.save(file_path)
    workbook.close()
    if cached_value is None:
        return
    temp_path = file_path.with_name(f'{file_path.stem}_tmp.xlsx')
    with ZipFile(file_path) as zip_file, \
            ZipFile(temp_path, 'w') as temp_zip:
        for item in zip_file.infolist():
            data = zip_file.read(item.filename)
            if item.filename == 'xl/worksheets/sheet1.xml':
                root = ET.fromstring(data)
                cell = root.find('.//sheet:c[@r="A1"]', _XML_NS)
                assert cell is not None
                value = cell.find('sheet:v', _XML_NS)
                assert value is not None
                value.text = str(cached_value)
                data = ET.tostring(root, encoding='utf-8',
                                   xml_declaration=False)
            temp_zip.writestr(item, data)
    temp_path.replace(file_path)


def create_update_workbook(file_path: Path) -> None:
    """Create the starting workbook used by the UPDATE mode test."""
    workbook = Workbook()
    worksheet = workbook.active
    assert isinstance(worksheet, Worksheet)
    worksheet['A1'] = 'old'
    worksheet['B1'] = 'row'
    workbook.save(file_path)
    workbook.close()


def inspect_updated_workbook(file_path: Path) -> None:
    """Check the workbook produced by the shared UPDATE mode case."""
    workbook = load_workbook(file_path, data_only=True)
    worksheet = workbook.active
    assert isinstance(worksheet, Worksheet)
    assert worksheet['A1'].value == 'old'
    assert worksheet['B1'].value == 'row'
    assert worksheet['A2'].value is None
    assert worksheet['B2'].value is None
    assert worksheet['A3'].value == 'new'
    assert worksheet['B3'].value == 'row'
    workbook.close()


def inspect_find_and_write_cells_workbook(
        file_path: Path, expect_highlight: bool = False) -> None:
    """Check exact cell writes after finding one row in the worksheet."""
    workbook = load_workbook(file_path)
    worksheet = workbook.active
    assert isinstance(worksheet, Worksheet)
    assert worksheet['B4'].value == 'Bob'
    assert worksheet['C4'].value is True
    if expect_highlight:
        assert worksheet['B4'].fill.fgColor.rgb == 'FFFFFF00'
        assert worksheet['C4'].fill.fgColor.rgb == 'FFFFFF00'
    workbook.close()
